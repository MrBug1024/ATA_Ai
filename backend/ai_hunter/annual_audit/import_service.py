"""Structured annual-audit imports triggered by the original chat upload flow.

Raw files remain in the platform object/evidence stores.  This module only
projects audit-relevant rows into the isolated MySQL annual-audit schema and
keeps a source locator on every projected row.
"""

from __future__ import annotations

import base64
import csv
import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable, Literal

from ai_hunter.app.services.minio_service import get_minio_service
from ai_hunter.app.settings import Settings, get_settings

from .engagement_repository import get_engagement
from .storage import mysql_connection, postgres_connection


DatasetType = Literal[
    "account_balance",
    "journal_entry",
    "receivable_item",
    "bank_transaction",
]


def _loads(value: Any) -> Any:
    """Decode JSON columns from either dict-row strings or native objects."""

    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return {}
    return value


@dataclass(frozen=True)
class TabularSheet:
    name: str
    rows: list[list[Any]]
    excel_datemode: int | None = None


_DATASET_LABELS: dict[DatasetType, str] = {
    "account_balance": "科目余额表",
    "journal_entry": "序时账/凭证分录",
    "receivable_item": "应收账款明细",
    "bank_transaction": "银行流水",
}

_ALIASES: dict[DatasetType, dict[str, tuple[str, ...]]] = {
    "account_balance": {
        "period_end": ("截止日期", "期末日期", "会计期间", "期间"),
        "account_code": ("科目编码", "科目代码", "科目编号", "账号编码"),
        "account_name": ("科目名称", "会计科目", "科目"),
        "opening_debit": ("期初借方", "期初借方余额"),
        "opening_credit": ("期初贷方", "期初贷方余额"),
        "period_debit": ("本期借方", "本期借方发生额", "借方发生额"),
        "period_credit": ("本期贷方", "本期贷方发生额", "贷方发生额"),
        "closing_debit": ("期末借方", "期末借方余额"),
        "closing_credit": ("期末贷方", "期末贷方余额"),
        "currency": ("币种", "货币"),
    },
    "journal_entry": {
        "voucher_date": ("凭证日期", "记账日期", "业务日期", "日期"),
        "voucher_no": ("凭证号", "凭证号数", "凭证编号", "记字号", "凭证号码"),
        "line_no": ("分录号", "行号", "序号"),
        "account_code": ("科目编码", "科目代码", "科目编号"),
        "account_name": ("科目名称", "会计科目", "科目"),
        "debit_amount": ("借方金额", "借方发生额", "借方"),
        "credit_amount": ("贷方金额", "贷方发生额", "贷方"),
        "counterparty": ("对方单位", "客商", "客户供应商", "往来单位", "对手方"),
        "description": ("摘要", "凭证摘要", "业务摘要", "说明"),
    },
    "receivable_item": {
        "customer_name": ("客户名称", "往来单位", "单位名称", "客户", "客商名称"),
        "document_no": ("单据号", "业务单号", "发票号", "凭证号", "编号"),
        "occurrence_date": ("发生日期", "入账日期", "凭证日期", "业务日期"),
        "due_date": ("到期日", "到期日期", "应收日期"),
        "balance": ("期末余额", "期末未审数", "审定数", "应收余额", "余额", "账面金额"),
        "currency": ("币种", "货币"),
        "is_related_party": ("是否关联方", "关联方", "关联关系"),
    },
    "bank_transaction": {
        "bank_account": ("银行账号", "账户", "账号", "本方账号", "开户账号"),
        "transaction_date": ("交易日期", "记账日期", "入账日期", "日期"),
        "amount": ("交易金额", "发生额", "金额"),
        "inflow_amount": ("收入金额", "转入金额", "贷方金额", "收入"),
        "outflow_amount": ("支出金额", "转出金额", "借方金额", "支出"),
        "direction": ("收支方向", "借贷标志", "方向", "交易方向"),
        "counterparty": ("对方户名", "对方单位", "对手方", "对方名称"),
        "transaction_ref": ("流水号", "交易流水号", "交易号", "参考号"),
        "description": ("摘要", "用途", "附言", "交易摘要", "说明"),
        "running_balance": ("账户余额", "交易后余额", "余额"),
    },
}

_REQUIRED_FIELDS: dict[DatasetType, tuple[set[str], ...]] = {
    "account_balance": (
        {"account_name", "closing_debit"},
        {"account_name", "closing_credit"},
        {"account_code", "period_debit", "period_credit"},
    ),
    "journal_entry": (
        {"voucher_date", "voucher_no", "account_name", "debit_amount"},
        {"voucher_date", "voucher_no", "account_name", "credit_amount"},
    ),
    "receivable_item": ({"customer_name", "balance"},),
    "bank_transaction": (
        {"transaction_date", "amount"},
        {"transaction_date", "inflow_amount"},
        {"transaction_date", "outflow_amount"},
    ),
}

_FILE_HINTS: dict[DatasetType, tuple[str, ...]] = {
    "account_balance": ("科目余额", "余额表", "trialbalance"),
    "journal_entry": ("序时账", "凭证", "分录", "journal"),
    "receivable_item": ("应收", "往来", "receivable"),
    "bank_transaction": ("银行流水", "账户流水", "bankstatement"),
}


def _header_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[\s_\-—（）()【】\[\]/\\.:：]+", "", text)


_ALIAS_INDEX: dict[DatasetType, dict[str, str]] = {
    dataset: {
        _header_key(alias): canonical
        for canonical, aliases in fields.items()
        for alias in (canonical, *aliases)
    }
    for dataset, fields in _ALIASES.items()
}


def _dataset_hint(text: str) -> DatasetType | None:
    normalized = _header_key(text)
    for dataset, hints in _FILE_HINTS.items():
        if any(_header_key(hint) in normalized for hint in hints):
            return dataset
    return None


def _column_mapping(dataset: DatasetType, header: list[Any]) -> dict[str, int]:
    index = _ALIAS_INDEX[dataset]
    mapping: dict[str, int] = {}
    for column_index, value in enumerate(header):
        key = _header_key(value)
        canonical = index.get(key)
        if canonical and canonical not in mapping:
            mapping[canonical] = column_index
    return mapping


def _mapping_is_usable(dataset: DatasetType, mapping: dict[str, int]) -> bool:
    fields = set(mapping)
    return any(required <= fields for required in _REQUIRED_FIELDS[dataset])


def _realign_merged_receivable_header(
    rows: list[list[Any]],
    *,
    header_index: int,
    mapping: dict[str, int],
) -> dict[str, int]:
    """Move a merged receivable header to the adjacent populated data column.

    Audit workpapers commonly merge ``往来单位`` across columns B:C while the
    actual customer name is stored in C.  Plain tabular readers only retain the
    header in B, so the schema is otherwise detected correctly but every data
    row appears to have an empty customer.  Only shift when the neighbouring
    header cell is blank; ordinary flat exports keep their adjacent header and
    are left untouched.
    """

    header = rows[header_index] if header_index < len(rows) else []
    aligned = dict(mapping)

    # Prefer the audited closing amount over an earlier merged ``期末未审数``
    # header when both are present in a workpaper.
    for column, value in enumerate(header):
        if _header_key(value) == _header_key("审定数"):
            aligned["balance"] = column
            break

    customer_column = aligned.get("customer_name")
    if customer_column is None or customer_column + 1 >= len(header):
        return aligned
    if header[customer_column + 1] not in (None, ""):
        return aligned

    sample_rows = rows[header_index + 1 : header_index + 201]

    def populated_count(column: int) -> int:
        return sum(
            1
            for row in sample_rows
            if column < len(row) and _text(row[column]) not in {"", "简称", "单位名称", "客户名称"}
        )

    candidates = [customer_column]
    for offset in (1, 2):
        candidate = customer_column + offset
        if candidate >= len(header) or header[candidate] not in (None, ""):
            break
        candidates.append(candidate)
    best_column = max(candidates, key=populated_count)
    if populated_count(best_column) <= populated_count(customer_column):
        return aligned
    return {**aligned, "customer_name": best_column}


def _local_data_row_count(
    dataset: DatasetType,
    rows: list[list[Any]],
    *,
    header_index: int,
    mapping: dict[str, int],
) -> int:
    """Estimate how many immediately following rows match the detected table."""

    required_groups = _REQUIRED_FIELDS[dataset]

    def has_value(row: list[Any], field: str) -> bool:
        column = mapping.get(field)
        return column is not None and column < len(row) and row[column] not in (None, "")

    return sum(
        1
        for row in rows[header_index + 1 : header_index + 26]
        if any(all(has_value(row, field) for field in required) for required in required_groups)
    )


def detect_sheet_schema(
    rows: list[list[Any]],
    *,
    source_hint: str = "",
) -> tuple[DatasetType, int, dict[str, int]] | None:
    """Detect a supported flat-table header in the first thirty rows."""

    hinted = _dataset_hint(source_hint)
    best: tuple[int, DatasetType, int, dict[str, int]] | None = None
    for row_index, row in enumerate(rows[:50]):
        for dataset in _ALIASES:
            mapping = _column_mapping(dataset, row)
            if not _mapping_is_usable(dataset, mapping):
                continue
            if dataset == "receivable_item":
                mapping = _realign_merged_receivable_header(
                    rows,
                    header_index=row_index,
                    mapping=mapping,
                )
            score = (
                len(mapping) * 10
                + (5 if dataset == hinted else 0)
                - row_index
                + _local_data_row_count(
                    dataset,
                    rows,
                    header_index=row_index,
                    mapping=mapping,
                )
                * 20
            )
            candidate = (score, dataset, row_index, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None:
        return None
    _, dataset, row_index, mapping = best
    return dataset, row_index, mapping


def _decimal(value: Any) -> Decimal:
    if value in (None, "", "-", "—", "/"):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = unicodedata.normalize("NFKC", str(value)).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[￥¥$,，\s]", "", text.strip("()"))
    if text.endswith("%"):
        text = text[:-1]
    try:
        result = Decimal(text or "0")
    except InvalidOperation as exc:
        raise ValueError(f"无法识别金额：{value!r}") from exc
    return -result if negative else result


def _date_value(value: Any, *, datemode: int | None = None) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and datemode is not None:
        epoch = datetime(1899, 12, 30) if datemode == 0 else datetime(1904, 1, 1)
        return (epoch + timedelta(days=float(value))).date()
    text = unicodedata.normalize("NFKC", str(value)).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y年%m月%d日",
        "%Y%m%d",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法识别日期：{value!r}")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _truthy(value: Any) -> bool:
    return _header_key(value) in {"1", "true", "yes", "y", "是", "关联方", "关联"}


def _row_value(row: list[Any], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _source_locator(
    *,
    source_ref: str,
    file_name: str,
    source_sha256: str = "",
    sheet_name: str,
    row_number: int,
    raw_row: list[Any],
) -> dict[str, Any]:
    preview = " | ".join(_text(value) for value in raw_row if _text(value))[:1000]
    return {
        "source_ref": source_ref,
        "file_name": file_name,
        "source_sha256": source_sha256,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "cell_range": f"A{row_number}:XFD{row_number}",
        "quote_text": preview,
    }


def normalize_sheet_rows(
    sheet: TabularSheet,
    *,
    source_ref: str,
    file_name: str,
    source_sha256: str = "",
    source_hint: str = "",
    default_period_end: date | None = None,
) -> tuple[DatasetType, list[dict[str, Any]]] | None:
    """Normalize one detected source sheet into canonical domain rows."""

    detected = detect_sheet_schema(sheet.rows, source_hint=f"{source_hint} {sheet.name} {file_name}")
    if detected is None:
        return None
    dataset, header_index, mapping = detected
    normalized_rows: list[dict[str, Any]] = []
    for row_index, raw_row in enumerate(sheet.rows[header_index + 1 :], start=header_index + 2):
        if not any(value not in (None, "") for value in raw_row):
            continue
        locator = _source_locator(
            source_ref=source_ref,
            file_name=file_name,
            source_sha256=source_sha256,
            sheet_name=sheet.name,
            row_number=row_index,
            raw_row=raw_row,
        )
        try:
            if dataset == "account_balance":
                account_name = _text(_row_value(raw_row, mapping, "account_name"))
                account_code = _text(_row_value(raw_row, mapping, "account_code"))
                if not account_name and not account_code:
                    continue
                normalized_rows.append(
                    {
                        "period_end": _date_value(
                            _row_value(raw_row, mapping, "period_end"),
                            datemode=sheet.excel_datemode,
                        )
                        or default_period_end,
                        "account_code": account_code or account_name,
                        "account_name": account_name or account_code,
                        "opening_debit": _decimal(_row_value(raw_row, mapping, "opening_debit")),
                        "opening_credit": _decimal(_row_value(raw_row, mapping, "opening_credit")),
                        "period_debit": _decimal(_row_value(raw_row, mapping, "period_debit")),
                        "period_credit": _decimal(_row_value(raw_row, mapping, "period_credit")),
                        "closing_debit": _decimal(_row_value(raw_row, mapping, "closing_debit")),
                        "closing_credit": _decimal(_row_value(raw_row, mapping, "closing_credit")),
                        "currency": _text(_row_value(raw_row, mapping, "currency")) or "CNY",
                        "source_locator_json": locator,
                    }
                )
            elif dataset == "journal_entry":
                voucher_no = _text(_row_value(raw_row, mapping, "voucher_no"))
                account_name = _text(_row_value(raw_row, mapping, "account_name"))
                voucher_date = _date_value(
                    _row_value(raw_row, mapping, "voucher_date"),
                    datemode=sheet.excel_datemode,
                )
                if not voucher_no or not account_name or voucher_date is None:
                    continue
                line_no_text = _text(_row_value(raw_row, mapping, "line_no"))
                normalized_rows.append(
                    {
                        "voucher_date": voucher_date,
                        "voucher_no": voucher_no,
                        "line_no": int(Decimal(line_no_text)) if line_no_text else len(normalized_rows) + 1,
                        "account_code": _text(_row_value(raw_row, mapping, "account_code")) or account_name,
                        "account_name": account_name,
                        "debit_amount": _decimal(_row_value(raw_row, mapping, "debit_amount")),
                        "credit_amount": _decimal(_row_value(raw_row, mapping, "credit_amount")),
                        "counterparty": _text(_row_value(raw_row, mapping, "counterparty")) or None,
                        "description": _text(_row_value(raw_row, mapping, "description")) or None,
                        "source_locator_json": locator,
                    }
                )
            elif dataset == "receivable_item":
                customer_name = _text(_row_value(raw_row, mapping, "customer_name"))
                if not customer_name:
                    continue
                normalized_rows.append(
                    {
                        "customer_name": customer_name,
                        "document_no": _text(_row_value(raw_row, mapping, "document_no")) or None,
                        "occurrence_date": _date_value(
                            _row_value(raw_row, mapping, "occurrence_date"),
                            datemode=sheet.excel_datemode,
                        ),
                        "due_date": _date_value(
                            _row_value(raw_row, mapping, "due_date"),
                            datemode=sheet.excel_datemode,
                        ),
                        "balance": _decimal(_row_value(raw_row, mapping, "balance")),
                        "currency": _text(_row_value(raw_row, mapping, "currency")) or "CNY",
                        "is_related_party": _truthy(_row_value(raw_row, mapping, "is_related_party")),
                        "source_locator_json": locator,
                    }
                )
            else:
                transaction_date = _date_value(
                    _row_value(raw_row, mapping, "transaction_date"),
                    datemode=sheet.excel_datemode,
                )
                if transaction_date is None:
                    continue
                amount_value = _row_value(raw_row, mapping, "amount")
                direction = _text(_row_value(raw_row, mapping, "direction"))
                if amount_value in (None, ""):
                    inflow = _decimal(_row_value(raw_row, mapping, "inflow_amount"))
                    outflow = _decimal(_row_value(raw_row, mapping, "outflow_amount"))
                    amount = inflow if inflow else outflow
                    direction = "in" if inflow else "out"
                else:
                    amount = _decimal(amount_value)
                    normalized_direction = _header_key(direction)
                    if normalized_direction in {"支出", "付款", "转出", "out", "debit", "借"}:
                        direction = "out"
                    elif normalized_direction in {"收入", "收款", "转入", "in", "credit", "贷"}:
                        direction = "in"
                    else:
                        direction = "out" if amount < 0 else "in"
                normalized_rows.append(
                    {
                        "bank_account": _text(_row_value(raw_row, mapping, "bank_account")) or "未标明账户",
                        "transaction_date": transaction_date,
                        "amount": abs(amount),
                        "direction": direction,
                        "counterparty": _text(_row_value(raw_row, mapping, "counterparty")) or None,
                        "transaction_ref": _text(_row_value(raw_row, mapping, "transaction_ref")) or None,
                        "description": _text(_row_value(raw_row, mapping, "description")) or None,
                        "running_balance": (
                            _decimal(_row_value(raw_row, mapping, "running_balance"))
                            if _row_value(raw_row, mapping, "running_balance") not in (None, "")
                            else None
                        ),
                        "source_locator_json": locator,
                    }
                )
        except (TypeError, ValueError, InvalidOperation):
            # Summary/footer/noise rows are expected in real accounting exports.
            continue
    if not normalized_rows:
        return None
    return dataset, normalized_rows


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_tabular_sheets(file_name: str, data: bytes) -> list[TabularSheet]:
    extension = Path(file_name).suffix.lower()
    if extension == ".csv":
        rows = [list(row) for row in csv.reader(StringIO(_decode_csv(data)))]
        return [TabularSheet(name="CSV", rows=rows)]
    if extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
        return [
            TabularSheet(name=sheet.title, rows=[list(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
    if extension == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        return [
            TabularSheet(
                name=sheet.name,
                rows=[sheet.row_values(index) for index in range(sheet.nrows)],
                excel_datemode=workbook.datemode,
            )
            for sheet in workbook.sheets()
        ]
    return []


def is_audit_workpaper_workbook(sheets: Iterable[TabularSheet]) -> bool:
    """Separate audit workpapers/templates from source accounting exports.

    Client workpaper packs contain transaction-like tables (dates, amounts,
    counterparties) but they are derived/output documents, not source facts.
    """

    names = {_header_key(sheet.name) for sheet in sheets}
    structural_markers = {
        "底稿目录",
        "审计程序",
        "审定表",
        "函证结果汇总表",
        "附注标准",
        "dxnsjtempsheet",
    }
    marker_count = sum(
        1
        for marker in structural_markers
        if any(marker in sheet_name for sheet_name in names)
    )
    return marker_count >= 2


def _file_bytes(file_item: dict[str, Any]) -> bytes:
    storage_ref = str(file_item.get("storage_ref") or file_item.get("content_ref") or "").strip()
    if storage_ref.startswith("minio://"):
        return get_minio_service().get_object_bytes(storage_ref)
    content = str(file_item.get("content") or "")
    if not content:
        raise ValueError("附件没有可读取的 storage_ref 或内联内容")
    if Path(str(file_item.get("name") or "")).suffix.lower() == ".csv":
        return content.encode("utf-8")
    if content.startswith("data:") and ";base64," in content:
        content = content.split(",", 1)[1]
    return base64.b64decode(content)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _persist_dataset(
    *,
    engagement_id: int,
    dataset: DatasetType,
    rows: list[dict[str, Any]],
    source_ref: str,
    source_sha256: str,
    file_name: str,
    created_by: str,
    settings: Settings,
) -> tuple[int, bool]:
    with mysql_connection(settings) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, row_count FROM annual_import_batch
                WHERE engagement_id = %s AND source_type = %s
                  AND source_sha256 = %s AND source_ref = %s AND status = 'completed'
                ORDER BY id DESC LIMIT 1
                """,
                (engagement_id, dataset, source_sha256, source_ref),
            )
            existing = cursor.fetchone()
            if existing:
                return int(existing["row_count"] or 0), True
            cursor.execute(
                """
                INSERT INTO annual_import_batch (
                  engagement_id, source_ref, source_type, source_sha256,
                  status, metadata_json, created_by
                ) VALUES (%s, %s, %s, %s, 'processing', %s, %s)
                """,
                (
                    engagement_id,
                    source_ref,
                    dataset,
                    source_sha256,
                    _json({"file_name": file_name, "dataset_label": _DATASET_LABELS[dataset]}),
                    created_by,
                ),
            )
            batch_id = int(cursor.lastrowid)

            if dataset == "account_balance":
                sql = """
                    INSERT INTO annual_account_balance (
                      engagement_id, import_batch_id, period_end, account_code, account_name,
                      opening_debit, opening_credit, period_debit, period_credit,
                      closing_debit, closing_credit, currency, source_locator_json,
                      source_file_id, source_page_id, source_chunk_id, locator_kind
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 'sheet_row')
                """
                values = [
                    (
                        engagement_id, batch_id, row["period_end"], row["account_code"], row["account_name"],
                        row["opening_debit"], row["opening_credit"], row["period_debit"], row["period_credit"],
                        row["closing_debit"], row["closing_credit"], row["currency"], _json(row["source_locator_json"]),
                    )
                    for row in rows
                    if row.get("period_end")
                ]
            elif dataset == "journal_entry":
                sql = """
                    INSERT INTO annual_journal_entry_line (
                      engagement_id, import_batch_id, voucher_date, voucher_no, line_no,
                      account_code, account_name, debit_amount, credit_amount,
                      counterparty, description, source_locator_json,
                      source_file_id, source_page_id, source_chunk_id, locator_kind
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 'sheet_row')
                """
                values = [
                    (
                        engagement_id, batch_id, row["voucher_date"], row["voucher_no"], row["line_no"],
                        row["account_code"], row["account_name"], row["debit_amount"], row["credit_amount"],
                        row["counterparty"], row["description"], _json(row["source_locator_json"]),
                    )
                    for row in rows
                ]
            elif dataset == "receivable_item":
                sql = """
                    INSERT INTO annual_receivable_item (
                      engagement_id, import_batch_id, customer_name, document_no,
                      occurrence_date, due_date, balance, currency, is_related_party,
                      source_locator_json, source_file_id, source_page_id, source_chunk_id, locator_kind
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 'sheet_row')
                """
                values = [
                    (
                        engagement_id, batch_id, row["customer_name"], row["document_no"],
                        row["occurrence_date"], row["due_date"], row["balance"], row["currency"],
                        row["is_related_party"], _json(row["source_locator_json"]),
                    )
                    for row in rows
                ]
            else:
                sql = """
                    INSERT INTO annual_bank_transaction (
                      engagement_id, import_batch_id, bank_account, transaction_date,
                      amount, direction, counterparty, transaction_ref, description,
                      running_balance, source_locator_json,
                      source_file_id, source_page_id, source_chunk_id, locator_kind
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 'sheet_row')
                """
                values = [
                    (
                        engagement_id, batch_id, row["bank_account"], row["transaction_date"],
                        row["amount"], row["direction"], row["counterparty"], row["transaction_ref"],
                        row["description"], row["running_balance"], _json(row["source_locator_json"]),
                    )
                    for row in rows
                ]
            if not values:
                raise ValueError(f"{_DATASET_LABELS[dataset]}没有可落库的有效数据行")
            cursor.executemany(sql, values)
            cursor.execute(
                """
                UPDATE annual_import_batch
                SET status = 'completed', row_count = %s, completed_at = NOW(6)
                WHERE id = %s
                """,
                (len(values), batch_id),
            )
        connection.commit()
    return len(values), False


def import_uploaded_files(
    *,
    engagement_id: int,
    files: Iterable[dict[str, Any]],
    actor: str = "ai_agent",
    settings: Settings | None = None,
) -> dict[str, Any]:
    """Project supported chat attachments into annual structured tables."""

    resolved = settings or get_settings()
    engagement = get_engagement(engagement_id, settings=resolved)
    imported: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    for file_item in files:
        file_name = str(file_item.get("name") or "uploaded-file")
        extension = Path(file_name).suffix.lower()
        if extension not in {".csv", ".xlsx", ".xlsm", ".xls"}:
            skipped.append({"file_name": file_name, "reason": "非结构化表格附件，继续走原 OCR/知识入库链路"})
            continue
        try:
            data = _file_bytes(file_item)
            source_sha256 = str(file_item.get("file_hash") or "") or hashlib.sha256(data).hexdigest()
            source_ref = str(file_item.get("storage_ref") or file_item.get("content_ref") or file_name)
            sheets = read_tabular_sheets(file_name, data)
            if is_audit_workpaper_workbook(sheets):
                skipped.append(
                    {
                        "file_name": file_name,
                        "reason": "识别为审计底稿/模板，仅进入原知识与模板链路，不写结构化事实表",
                    }
                )
                continue
            matched = 0
            for sheet in sheets:
                normalized = normalize_sheet_rows(
                    sheet,
                    source_ref=source_ref,
                    file_name=file_name,
                    source_sha256=source_sha256,
                    source_hint=str(file_item.get("doc_category") or ""),
                    default_period_end=engagement["period_end"],
                )
                if normalized is None:
                    continue
                dataset, rows = normalized
                row_count, deduplicated = _persist_dataset(
                    engagement_id=engagement_id,
                    dataset=dataset,
                    rows=rows,
                    source_ref=f"{source_ref}#sheet={sheet.name}",
                    source_sha256=source_sha256,
                    file_name=file_name,
                    created_by=actor or "ai_agent",
                    settings=resolved,
                )
                imported.append(
                    {
                        "file_name": file_name,
                        "sheet_name": sheet.name,
                        "dataset": dataset,
                        "dataset_label": _DATASET_LABELS[dataset],
                        "row_count": row_count,
                        "deduplicated": deduplicated,
                    }
                )
                matched += 1
            if matched == 0:
                skipped.append(
                    {
                        "file_name": file_name,
                        "reason": "未识别到科目余额、序时账、应收明细或银行流水标准表头",
                    }
                )
        except Exception as exc:  # one bad workbook must not block original OCR ingestion
            errors.append({"file_name": file_name, "error": str(exc)[:500]})
    row_count = sum(int(item["row_count"]) for item in imported if not item["deduplicated"])
    return {
        "status": "completed" if not errors else "partial",
        "engagement_id": engagement_id,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "new_row_count": row_count,
        "recognized_datasets": sorted({item["dataset"] for item in imported}),
    }


_SHEET_HEADER_RE = re.compile(r"^\s*##\s*工作表\s*[：:]\s*(.+?)\s*$", re.MULTILINE)
_SHEET_ROW_RE = re.compile(r"\[第\s*(\d+)\s*行\]")


def _canonical_source_key(value: Any) -> str:
    """Normalize a source identity for deterministic equality matching."""

    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", _text(value))).strip().casefold()


def _positive_int(value: Any) -> int:
    try:
        return max(0, int(value or 0))
    except (TypeError, ValueError):
        return 0


def _unique_rows(rows: Iterable[dict[str, Any]], *, identity_field: str) -> list[dict[str, Any]]:
    """Deduplicate candidates by a durable identifier before testing uniqueness."""

    unique: dict[str, dict[str, Any]] = {}
    for row in rows:
        identity = str(row.get(identity_field) or "").strip()
        if identity:
            unique[identity] = row
    return [unique[key] for key in sorted(unique)]


def _page_sheet_keys(page: dict[str, Any]) -> set[str]:
    """Return all worksheet names carried by one source-page batch record."""

    values: list[Any] = [page.get("sheet_name")]
    sheet_names = page.get("sheet_names")
    if isinstance(sheet_names, list):
        values.extend(sheet_names)
    page_text = str(page.get("page_text") or "")
    values.extend(match.group(1) for match in _SHEET_HEADER_RE.finditer(page_text))
    return {key for value in values if (key := _canonical_source_key(value))}


def _select_source_file(
    *,
    locator: dict[str, Any],
    files_by_sha256: dict[str, list[dict[str, Any]]],
    files_by_name: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    """Select one canonical platform file, never a positional fallback."""

    source_sha256 = _canonical_source_key(locator.get("source_sha256"))
    if source_sha256:
        candidates = _unique_rows(files_by_sha256.get(source_sha256, []), identity_field="id")
        if len(candidates) == 1:
            return candidates[0], ""
        return None, "source_file_not_found" if not candidates else "source_file_ambiguous"

    file_name = _canonical_source_key(locator.get("file_name"))
    if not file_name:
        return None, "source_file_missing"
    candidates = _unique_rows(files_by_name.get(file_name, []), identity_field="id")
    if len(candidates) == 1:
        return candidates[0], ""
    return None, "source_file_not_found" if not candidates else "source_file_ambiguous"


def _select_source_page(
    *,
    pages: list[dict[str, Any]],
    sheet_name: str,
    allow_unique_page_without_sheet: bool = False,
) -> tuple[dict[str, Any] | None, str]:
    """Select exactly one source page for a worksheet."""

    candidates = _unique_rows(pages, identity_field="id")
    expected_sheet = _canonical_source_key(sheet_name)
    if expected_sheet:
        sheet_matches = [page for page in candidates if expected_sheet in _page_sheet_keys(page)]
        if sheet_matches:
            candidates = sheet_matches
        elif not (allow_unique_page_without_sheet and len(candidates) == 1):
            return None, "sheet_not_found"
    if len(candidates) == 1:
        return candidates[0], ""
    return None, "page_not_found" if not candidates else "page_ambiguous"


def _chunk_row_bounds(chunk: dict[str, Any]) -> tuple[int, int]:
    """Resolve spreadsheet row bounds from explicit metadata or source text."""

    metadata = _loads(chunk.get("metadata")) or {}
    metadata = metadata if isinstance(metadata, dict) else {}
    row_start = _positive_int(metadata.get("row_start"))
    row_end = _positive_int(metadata.get("row_end"))
    if row_start > 0 and row_end >= row_start:
        return row_start, row_end

    rows = [int(match.group(1)) for match in _SHEET_ROW_RE.finditer(str(chunk.get("chunk_text") or ""))]
    if not rows:
        return 0, 0
    return min(rows), max(rows)


def _chunk_contains_row(chunk: dict[str, Any], row_number: int) -> bool:
    row_start, row_end = _chunk_row_bounds(chunk)
    return row_start > 0 and row_start <= row_number <= row_end


def _chunk_contains_quote(chunk: dict[str, Any], quote: str) -> bool:
    expected = _canonical_source_key(quote)
    return bool(expected and expected in _canonical_source_key(chunk.get("chunk_text")))


def _select_source_chunk(
    *,
    chunks: list[dict[str, Any]],
    row_number: int,
    quote_text: str,
) -> tuple[dict[str, Any] | None, str]:
    """Select exactly one page chunk using a spreadsheet row or exact quote."""

    candidates = _unique_rows(chunks, identity_field="chunk_id")
    if not candidates:
        return None, "chunk_not_found"

    quote_matches = [chunk for chunk in candidates if _chunk_contains_quote(chunk, quote_text)]
    if row_number > 0:
        row_matches = [chunk for chunk in candidates if _chunk_contains_row(chunk, row_number)]
        if len(row_matches) == 1:
            return row_matches[0], ""
        if len(row_matches) > 1:
            row_quote_matches = [chunk for chunk in row_matches if _chunk_contains_quote(chunk, quote_text)]
            if len(row_quote_matches) == 1:
                return row_quote_matches[0], ""
            return None, "row_range_ambiguous"
        if len(quote_matches) == 1:
            return quote_matches[0], ""
        if len(quote_matches) > 1:
            return None, "quote_ambiguous"
        return None, "row_range_not_found"

    if len(quote_matches) == 1:
        return quote_matches[0], ""
    if len(quote_matches) > 1:
        return None, "quote_ambiguous"
    return None, "chunk_not_found" if quote_text else "row_and_quote_missing"


def bind_structured_source_refs(
    *,
    engagement_id: int,
    chunk_batch: dict[str, Any],
    settings: Settings | None = None,
) -> dict[str, Any]:
    """Bind MySQL structured rows to real PostgreSQL source anchors.

    Structured import intentionally happens before the normal graph ingest.
    This second pass runs after ``load_chunks`` has created source files,
    pages and chunks.  It replaces the old ``annual:*`` synthetic references
    with real platform IDs and preserves spreadsheet row/sheet coordinates.
    """

    resolved = settings or get_settings()
    if int(engagement_id or 0) <= 0:
        return {"bound_count": 0, "unbound_count": 0, "status": "no_engagement"}

    file_rows = [row for row in chunk_batch.get("files", []) if isinstance(row, dict)]
    page_rows = [row for row in chunk_batch.get("pages", []) if isinstance(row, dict)]
    chunk_rows = [row for row in chunk_batch.get("chunks", []) if isinstance(row, dict)]
    if not file_rows:
        return {"bound_count": 0, "unbound_count": 0, "status": "no_source_files"}

    platform_files_by_sha256: dict[str, list[dict[str, Any]]] = {}
    platform_files_by_name: dict[str, list[dict[str, Any]]] = {}
    platform_file_ids: set[int] = set()
    for row in file_rows:
        file_id = _positive_int(row.get("id"))
        if file_id <= 0:
            continue
        normalized = {
            "id": file_id,
            "file_name": str(row.get("file_name") or ""),
            "file_sha256": str(row.get("file_sha256") or ""),
            "content_type": str(row.get("content_type") or ""),
            "storage_ref": str(row.get("storage_ref") or ""),
        }
        platform_file_ids.add(file_id)
        source_sha256 = _canonical_source_key(normalized["file_sha256"])
        if source_sha256:
            platform_files_by_sha256.setdefault(source_sha256, []).append(normalized)
        file_name = _canonical_source_key(normalized["file_name"])
        if file_name:
            platform_files_by_name.setdefault(file_name, []).append(normalized)

    page_by_file: dict[int, list[dict[str, Any]]] = {}
    for row in page_rows:
        page_by_file.setdefault(_positive_int(row.get("file_id")), []).append(row)
    chunk_by_page: dict[int, list[dict[str, Any]]] = {}
    for row in chunk_rows:
        chunk_by_page.setdefault(_positive_int(row.get("page_id")), []).append(row)

    table_map = {
        "account_balance": "annual_account_balance",
        "journal_entry": "annual_journal_entry_line",
        "receivable_item": "annual_receivable_item",
        "bank_transaction": "annual_bank_transaction",
    }
    bound = 0
    unbound = 0
    unbound_reasons: dict[str, int] = {}
    unbound_rows: list[dict[str, Any]] = []

    def record_unbound(
        *,
        dataset: str,
        source_row: dict[str, Any],
        locator: dict[str, Any],
        reason: str,
    ) -> None:
        nonlocal unbound
        unbound += 1
        unbound_reasons[reason] = unbound_reasons.get(reason, 0) + 1
        if len(unbound_rows) >= 100:
            return
        unbound_rows.append(
            {
                "dataset": dataset,
                "record_id": _positive_int(source_row.get("id")),
                "reason": reason,
                "file_name": str(locator.get("file_name") or ""),
                "sheet_name": str(locator.get("sheet_name") or ""),
                "row_number": _positive_int(locator.get("row_number") or locator.get("row_start")),
            }
        )

    with mysql_connection(resolved) as mysql_conn, postgres_connection(resolved) as pg_conn:
        with mysql_conn.cursor() as mysql_cursor, pg_conn.cursor() as pg_cursor:
            for dataset, table_name in table_map.items():
                mysql_cursor.execute(
                    f"""
                    SELECT id, source_locator_json
                    FROM {table_name}
                    WHERE engagement_id = %s
                      AND (
                        source_file_id IS NULL
                        OR source_page_id IS NULL
                        OR source_chunk_id IS NULL
                        OR source_chunk_id = ''
                      )
                    ORDER BY id
                    """,
                    (engagement_id,),
                )
                rows = list(mysql_cursor.fetchall())
                for row in rows:
                    locator = _loads(row.get("source_locator_json")) or {}
                    if not isinstance(locator, dict):
                        record_unbound(
                            dataset=dataset,
                            source_row=row,
                            locator={},
                            reason="invalid_source_locator",
                        )
                        continue

                    file_row, reason = _select_source_file(
                        locator=locator,
                        files_by_sha256=platform_files_by_sha256,
                        files_by_name=platform_files_by_name,
                    )
                    if file_row is None:
                        record_unbound(
                            dataset=dataset,
                            source_row=row,
                            locator=locator,
                            reason=reason,
                        )
                        continue

                    file_id = _positive_int(file_row["id"])
                    candidate_pages = page_by_file.get(file_id, [])
                    sheet_name = str(locator.get("sheet_name") or "")
                    page, reason = _select_source_page(
                        pages=candidate_pages,
                        sheet_name=sheet_name,
                        allow_unique_page_without_sheet=(
                            str(file_row["file_name"]).lower().endswith(".csv")
                            or str(file_row["content_type"]).lower().startswith("text/csv")
                        ),
                    )
                    if page is None:
                        record_unbound(
                            dataset=dataset,
                            source_row=row,
                            locator=locator,
                            reason=reason,
                        )
                        continue

                    page_id = _positive_int(page.get("id"))
                    if page_id <= 0:
                        record_unbound(
                            dataset=dataset,
                            source_row=row,
                            locator=locator,
                            reason="page_not_found",
                        )
                        continue
                    quote = str(locator.get("quote_text") or "").strip()
                    chunk, reason = _select_source_chunk(
                        chunks=chunk_by_page.get(page_id, []),
                        row_number=_positive_int(locator.get("row_number") or locator.get("row_start")),
                        quote_text=quote,
                    )
                    chunk_id = str(chunk.get("chunk_id") or "").strip() if chunk else ""
                    if not chunk_id:
                        record_unbound(
                            dataset=dataset,
                            source_row=row,
                            locator=locator,
                            reason=reason or "chunk_not_found",
                        )
                        continue

                    chunk_row_start, chunk_row_end = _chunk_row_bounds(chunk)
                    row_start = (
                        _positive_int(locator.get("row_start"))
                        or _positive_int(locator.get("row_number"))
                        or chunk_row_start
                    )
                    row_end = _positive_int(locator.get("row_end")) or row_start or chunk_row_end
                    if row_start > 0 and row_end < row_start:
                        row_end = row_start
                    cell_range = str(locator.get("cell_range") or "").strip()
                    if not cell_range and row_start > 0:
                        cell_range = f"A{row_start}:XFD{row_end}"

                    pg_cursor.execute(
                        """
                        SELECT page_image_ref, page_width, page_height
                        FROM public.source_page
                        WHERE id = %s
                        """,
                        (page_id,),
                    )
                    page_meta = dict(pg_cursor.fetchone() or {})
                    locator.update(
                        {
                            "domain_code": "annual_audit",
                            "project_id": int(engagement_id),
                            "source_file_id": file_id,
                            "source_page_id": page_id,
                            "source_chunk_id": chunk_id,
                            "locator_kind": (
                                "csv_row"
                                if str(file_row["file_name"]).lower().endswith(".csv")
                                else "sheet_row"
                            ),
                            "page_no": int(page.get("page_no") or 0),
                            "row_start": row_start,
                            "row_end": row_end,
                            "cell_range": cell_range,
                            "source_chunk_row_start": chunk_row_start,
                            "source_chunk_row_end": chunk_row_end,
                            "content_type": file_row["content_type"],
                            "source_file_ref": file_row["storage_ref"],
                            "page_image_ref": str(page_meta.get("page_image_ref") or ""),
                            "page_width": int(page_meta.get("page_width") or 0),
                            "page_height": int(page_meta.get("page_height") or 0),
                            "preview_available": bool(chunk_id),
                        }
                    )
                    mysql_cursor.execute(
                        f"""
                        UPDATE {table_name}
                        SET source_file_id = %s,
                            source_page_id = %s,
                            source_chunk_id = %s,
                            locator_kind = %s,
                            source_locator_json = %s
                        WHERE id = %s AND engagement_id = %s
                        """,
                        (
                            file_id,
                            page_id,
                            chunk_id,
                            str(locator["locator_kind"]),
                            _json(locator),
                            int(row["id"]),
                            engagement_id,
                        ),
                    )
                    bound += 1
        mysql_conn.commit()
        pg_conn.commit()
    return {
        "status": "completed" if unbound == 0 else "partial",
        "bound_count": bound,
        "unbound_count": unbound,
        "source_file_count": len(platform_file_ids),
        "unbound_reasons": dict(sorted(unbound_reasons.items())),
        "unbound_rows": unbound_rows,
        "unbound_rows_truncated": unbound > len(unbound_rows),
    }


__all__ = [
    "TabularSheet",
    "detect_sheet_schema",
    "import_uploaded_files",
    "bind_structured_source_refs",
    "is_audit_workpaper_workbook",
    "normalize_sheet_rows",
    "read_tabular_sheets",
]
