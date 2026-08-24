"""Generate the confirmed annual-audit attachment package.

The full-audit turn produces a review draft.  This module is intentionally a
separate second step: it is called only after the user confirms the result and
then confirms generation.  Every package stores the active template snapshot
used for each attachment, which makes a later template activation unable to
silently change an earlier delivery.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

from ai_hunter.app.settings import Settings, get_settings
from ai_hunter.platform_core import ArtifactRef

from .artifact_service import (
    REPORT_CONTENT_TYPES,
    build_docx,
    build_report_xlsx,
    build_workpaper_xlsx,
)
from .storage import mysql_connection
from .generic_template_repository import get_active_template_catalog, template_version_ref


DEFAULT_ATTACHMENT_TYPES = (
    "annual_report",
    "financial_statements",
    "notes",
    "management_letter",
)


def _json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _template_title(template: dict[str, Any], fallback: str) -> str:
    content = template.get("content") or {}
    return str(content.get("title") or fallback)


def _load_latest_report(engagement_id: int, settings: Settings) -> dict[str, Any]:
    with mysql_connection(settings) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, report_version, template_version, status, fact_snapshot_json
                FROM audit_report
                WHERE engagement_id = %s AND report_type = 'annual_audit_draft'
                ORDER BY report_version DESC LIMIT 1
                """,
                (engagement_id,),
            )
            row = dict(cursor.fetchone() or {})
    if not row:
        raise ValueError("当前项目尚未生成审计结果，不能生成附件")
    snapshot = row.get("fact_snapshot_json")
    if isinstance(snapshot, str):
        try:
            snapshot = json.loads(snapshot)
        except json.JSONDecodeError:
            snapshot = {}
    row["snapshot"] = snapshot if isinstance(snapshot, dict) else {}
    return row


def _next_package_version(engagement_id: int, settings: Settings) -> int:
    with mysql_connection(settings) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COALESCE(MAX(package_version), 0) AS latest FROM annual_audit_attachment_package WHERE engagement_id = %s",
                (engagement_id,),
            )
            return int((cursor.fetchone() or {}).get("latest") or 0) + 1


def _build_financial_statements_xlsx(
    *,
    snapshot: dict[str, Any],
    template: dict[str, Any],
    package_version: int,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet_name = str((template.get("content") or {}).get("sheet_name") or "财务报表")[:31]
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append([_template_title(template, "年度财务报表"), f"附件包 v{package_version}"])
    sheet.append(["模板版本", template_version_ref(template)])
    sheet.append([])
    sheet.append(["项目", "数值"])
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    readiness = snapshot.get("readiness") or {}
    counts = readiness.get("counts") or {}
    rows = (
        ("被审计单位", (snapshot.get("engagement") or {}).get("entity_name", "")),
        ("科目余额行数", counts.get("account_balance_rows", 0)),
        ("序时账/凭证明细行数", counts.get("journal_entry_rows", 0)),
        ("应收账款明细行数", counts.get("receivable_rows", 0)),
        ("银行流水行数", counts.get("bank_transaction_rows", 0)),
        ("收入净额", ((snapshot.get("sales_receivables") or {}).get("revenue") or {}).get("net_revenue", 0)),
        ("应收账款余额", ((snapshot.get("sales_receivables") or {}).get("receivables") or {}).get("total_balance", 0)),
        ("银行流入", (snapshot.get("cash_and_bank") or {}).get("total_inflow", 0)),
        ("银行流出", (snapshot.get("cash_and_bank") or {}).get("total_outflow", 0)),
    )
    for key, value in rows:
        sheet.append([key, value])
    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 48
    output = __import__("io").BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_notes_docx(*, report_text: str, template: dict[str, Any], package_version: int) -> bytes:
    sections = (template.get("content") or {}).get("sections") or []
    section_text = "\n".join(f"{index}. {name}" for index, name in enumerate(sections, start=1))
    body = (
        f"附件包版本：{package_version}\n"
        f"模板版本：{template_version_ref(template)}\n\n"
        f"附注章节模板：\n{section_text or '按项目组确认的附注清单编制'}\n\n"
        "本附注为基于当前审计结果形成的交付草稿，正式发布前须由项目组核对报表数字、会计政策和披露完整性。\n\n"
        f"审计结果摘要：\n{report_text[:12000]}"
    )
    return build_docx(title=_template_title(template, "财务报表附注"), report_text=body)


def _build_management_letter_docx(
    *,
    report_text: str,
    template: dict[str, Any],
    package_version: int,
) -> bytes:
    body = (
        f"附件包版本：{package_version}\n"
        f"模板版本：{template_version_ref(template)}\n\n"
        "以下内容根据当前已确认的审计结果整理，项目组应在正式发送前补充管理层沟通记录、责任部门和整改期限。\n\n"
        f"审计发现与建议基础：\n{report_text[:12000]}"
    )
    return build_docx(title=_template_title(template, "管理建议书"), report_text=body)


def _persist_package(
    *,
    engagement_id: int,
    package_version: int,
    status: str,
    template_snapshot: dict[str, Any],
    artifacts: list[dict[str, Any]],
    created_by: str,
    settings: Settings,
) -> int:
    with mysql_connection(settings) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO annual_audit_attachment_package
                  (engagement_id, package_version, status, template_snapshot_json,
                   artifact_refs_json, created_by)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    engagement_id,
                    package_version,
                    status,
                    json.dumps(template_snapshot, ensure_ascii=False, default=_json_default),
                    json.dumps(artifacts, ensure_ascii=False, default=_json_default),
                    created_by or "ai_agent",
                ),
            )
            package_id = int(cursor.lastrowid)
        connection.commit()
    return package_id


def generate_annual_attachment_package(
    engagement_id: int,
    *,
    created_by: str = "ai_agent",
    requested_types: list[str] | None = None,
    settings: Settings | None = None,
) -> dict[str, Any]:
    """Generate all confirmed annual-audit attachments from active templates."""

    # The file-backed renderer is the source of truth.  Keep the legacy
    # programmatic implementation below for compatibility with old imports,
    # but do not use it for confirmed deliveries.
    from .file_attachment_service import generate_annual_attachment_package as render_from_files

    return render_from_files(
        engagement_id,
        created_by=created_by,
        requested_types=requested_types,
        settings=settings,
    )

    resolved = settings or get_settings()
    selected_types = tuple(requested_types or DEFAULT_ATTACHMENT_TYPES)
    unknown = [item for item in selected_types if item not in DEFAULT_ATTACHMENT_TYPES]
    if unknown:
        raise ValueError(f"不支持的附件类型：{', '.join(unknown)}")

    # Reuse the deterministic report result as the source snapshot.  It is
    # generation-key idempotent and does not replace an already equal draft.
    from .report_service import generate_annual_report_draft

    generated = generate_annual_report_draft(
        engagement_id,
        recompute=False,
        created_by=created_by or "ai_agent",
        settings=resolved,
    )
    report_text = str(generated.get("report_text") or "")
    latest_report = _load_latest_report(engagement_id, resolved)
    snapshot = dict(latest_report.get("snapshot") or {})
    snapshot.setdefault("engagement", {})
    # The report generator has the authoritative text; the database snapshot
    # intentionally contains facts and is not used as a prose source.
    catalog = get_active_template_catalog(settings=resolved)
    package_version = _next_package_version(engagement_id, resolved)
    template_snapshot = {
        template_type: {
            "template_code": str(template.get("template_code") or ""),
            "template_type": template_type,
            "version_no": int(template.get("version_no") or 0),
            "version_label": template_version_ref(template),
            "content_hash": str(template.get("content_hash") or ""),
        }
        for template_type, template in catalog.items()
        if template_type in selected_types
    }

    workpapers = list((generated.get("artifacts") or {}).get("workpapers") or [])
    payloads: list[tuple[str, str, bytes, str, dict[str, Any]]] = []
    if "annual_report" in selected_types:
        report_template = catalog["annual_report"]
        payloads.append(
            (
                f"annual-audit-report-{engagement_id}-package-v{package_version}.docx",
                REPORT_CONTENT_TYPES["docx"],
                build_docx(
                    title=_template_title(report_template, "年度财务报表审计报告"),
                    report_text=report_text,
                ),
                "annual_report",
                report_template,
            )
        )
    if "financial_statements" in selected_types:
        payloads.append(
            (
                f"financial-statements-{engagement_id}-package-v{package_version}.xlsx",
                REPORT_CONTENT_TYPES["xlsx"],
                _build_financial_statements_xlsx(
                    snapshot=snapshot,
                    template=catalog["financial_statements"],
                    package_version=package_version,
                ),
                "financial_statements",
                catalog["financial_statements"],
            )
        )
    if "notes" in selected_types:
        payloads.append(
            (
                f"audit-notes-{engagement_id}-package-v{package_version}.docx",
                REPORT_CONTENT_TYPES["docx"],
                _build_notes_docx(
                    report_text=report_text,
                    template=catalog["notes"],
                    package_version=package_version,
                ),
                "notes",
                catalog["notes"],
            )
        )
    if "management_letter" in selected_types:
        payloads.append(
            (
                f"management-letter-{engagement_id}-package-v{package_version}.docx",
                REPORT_CONTENT_TYPES["docx"],
                _build_management_letter_docx(
                    report_text=report_text,
                    template=catalog["management_letter"],
                    package_version=package_version,
                ),
                "management_letter",
                catalog["management_letter"],
            )
        )
    if "audit_workpaper" in selected_types:
        workpaper_template = catalog["audit_workpaper"]
        for workpaper in workpapers:
            code = str(workpaper.get("code") or "workpaper")
            facts = (snapshot.get("workpaper_facts") or {}).get(code) or {}
            payloads.append(
                (
                    f"audit-workpaper-{code}-{engagement_id}-package-v{package_version}.xlsx",
                    REPORT_CONTENT_TYPES["xlsx"],
                    build_workpaper_xlsx(
                        code=code,
                        name=str(workpaper.get("name") or code),
                        facts=facts,
                        version=package_version,
                        title=_template_title(workpaper_template, "审计工作底稿"),
                        sheet_name=str((workpaper_template.get("content") or {}).get("sheet_name") or "底稿"),
                    ),
                    f"audit_workpaper_{code}",
                    workpaper_template,
                )
            )

    from .artifact_service import get_minio_service

    service = get_minio_service()
    published: list[ArtifactRef] = []
    errors: list[str] = []
    for file_name, content_type, data, artifact_type, template in payloads:
        try:
            uploaded = service.upload_artifact(
                project_id=engagement_id,
                file_name=file_name,
                content_type=content_type,
                file_bytes=data,
            )
            published.append(
                ArtifactRef(
                    artifact_type=artifact_type,
                    template_version=template_version_ref(template),
                    version=package_version,
                    storage_ref=uploaded.storage_ref,
                    content_type=content_type,
                    file_name=file_name,
                    status="draft",
                )
            )
        except Exception as exc:
            errors.append(f"{file_name}: {str(exc)[:240]}")

    artifact_dicts = [item.to_dict() for item in published]
    status = "draft_saved" if not errors else ("partial" if published else "failed")
    package_id = _persist_package(
        engagement_id=engagement_id,
        package_version=package_version,
        status=status,
        template_snapshot=template_snapshot,
        artifacts=artifact_dicts,
        created_by=created_by,
        settings=resolved,
    )
    return {
        "package_id": package_id,
        "package_version": package_version,
        "status": status,
        "template_snapshot": template_snapshot,
        "artifacts": artifact_dicts,
        "errors": errors,
    }


__all__ = ["DEFAULT_ATTACHMENT_TYPES", "generate_annual_attachment_package"]
