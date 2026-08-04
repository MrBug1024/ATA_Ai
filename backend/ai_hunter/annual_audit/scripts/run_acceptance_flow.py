"""Run one isolated, end-to-end annual-audit acceptance flow.

The script creates only a new annual-audit engagement, uploads an in-memory
workbook to the annual MinIO bucket, runs the real LangGraph ingestion and
analysis path, resolves evidence, and publishes the report/workpaper artifacts.
It never reads or writes the legacy NPA project.
"""

from __future__ import annotations

import base64
import hashlib
import io
import json
import time

from openpyxl import Workbook

from ai_hunter.annual_audit.engagement_repository import create_engagement
from ai_hunter.annual_audit.storage import mysql_connection
from ai_hunter.app.graph.main import build_audit_orchestrator_graph
from ai_hunter.app.services.minio_service import get_minio_service


def _build_acceptance_workbook() -> bytes:
    workbook = Workbook()
    account = workbook.active
    account.title = "科目余额表"
    account.append(["科目编码", "科目名称", "本期借方发生额", "本期贷方发生额", "期末借方余额", "期末贷方余额"])
    account.append(["1002", "银行存款", 2500000, 500000, 2000000, 0])
    account.append(["1122", "应收账款", 500000, 0, 500000, 0])
    account.append(["6001", "主营业务收入", 0, 3000000, 0, 3000000])

    journal = workbook.create_sheet("序时账")
    journal.append(["凭证日期", "凭证号", "分录号", "科目编码", "科目名称", "借方金额", "贷方金额", "摘要"])
    journal.append(["2025-12-31", "JV-001", 1, "1122", "应收账款", 500000, 0, "确认收入"])
    journal.append(["2025-12-31", "JV-001", 2, "6001", "主营业务收入", 0, 500000, "确认收入"])
    journal.append(["2025-12-30", "JV-002", 1, "1002", "银行存款", 2000000, 0, "大额收款"])

    receivable = workbook.create_sheet("应收账款明细")
    receivable.append(["客户名称", "单据号", "发生日期", "到期日", "期末余额", "是否关联方"])
    receivable.append(["客户甲", "INV-001", "2024-01-01", "2024-03-01", 300000, "否"])
    receivable.append(["客户乙", "INV-002", "2025-12-20", "2026-01-30", 200000, "是"])

    bank = workbook.create_sheet("银行流水")
    bank.append(["交易日期", "银行账号", "收入金额", "支出金额", "对方户名", "流水号"])
    bank.append(["2025-12-30", "6222", 2000000, 0, "客户甲", "T-001"])
    bank.append(["2025-12-31", "6222", 0, 1500000, "供应商乙", "T-002"])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def main() -> None:
    stamp = str(int(time.time()))
    entity_name = f"ATA验收企业-{stamp}"
    engagement = create_engagement(
        {
            "case_name": f"ATA年审完整流程验收-{stamp}",
            "entity_name": entity_name,
            "fiscal_year": 2025,
            "company_id": f"ata-acceptance-{stamp}",
            "owner_id": "acceptance-runner",
            "created_by": "acceptance-runner",
        }
    )
    case_id = int(engagement["case_id"])
    file_name = f"ATA-annual-acceptance-{stamp}.xlsx"
    file_bytes = _build_acceptance_workbook()
    upload = get_minio_service().upload_raw_file(
        case_id=case_id,
        entity_id=0,
        file_name=file_name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_bytes=file_bytes,
        entity_name=entity_name,
    )
    file_item = {
        "name": file_name,
        "type": "document",
        "extension": ".xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content": base64.b64encode(file_bytes).decode("ascii"),
        "file_hash": hashlib.sha256(file_bytes).hexdigest(),
        "file_size": len(file_bytes),
        "storage_ref": upload.storage_ref,
        "storage_provider": upload.storage_provider,
        "storage_bucket": upload.storage_bucket,
        "storage_key": upload.storage_key,
        "doc_category": "trial_balance",
    }

    thread_id = f"annual-acceptance-{stamp}"
    graph = build_audit_orchestrator_graph()
    result = graph.invoke(
        {
            "thread_id": thread_id,
            "query": "执行完整年审并生成报告",
            "current_case_id": case_id,
            "operator_id": "acceptance-runner",
            "operator_name": "acceptance-runner",
            "uploaded_files": [file_item],
        },
        config={"configurable": {"thread_id": thread_id}},
    )

    with mysql_connection() as connection, connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT COUNT(*) AS total_rows,
                   SUM(source_file_id IS NOT NULL) AS bound_files,
                   SUM(source_chunk_id IS NOT NULL AND source_chunk_id <> '') AS bound_chunks
            FROM (
              SELECT source_file_id, source_chunk_id FROM annual_account_balance WHERE engagement_id = %s
              UNION ALL SELECT source_file_id, source_chunk_id FROM annual_journal_entry_line WHERE engagement_id = %s
              UNION ALL SELECT source_file_id, source_chunk_id FROM annual_receivable_item WHERE engagement_id = %s
              UNION ALL SELECT source_file_id, source_chunk_id FROM annual_bank_transaction WHERE engagement_id = %s
            ) source_rows
            """,
            (case_id, case_id, case_id, case_id),
        )
        binding = dict(cursor.fetchone() or {})

    artifacts = result.get("artifacts") or {}
    trace_items = result.get("trace_items") or []
    evidence_count = sum(len(item.get("evidences") or []) for item in trace_items if isinstance(item, dict))
    report_ref = str(result.get("final_report_ref") or "")
    print(
        json.dumps(
            {
                "case_id": case_id,
                "thread_id": thread_id,
                "intent": result.get("intent"),
                "records_inserted": result.get("records_inserted"),
                "annual_import_summary": result.get("annual_import_summary"),
                "annual_evidence_binding_summary": result.get("annual_evidence_binding_summary"),
                "structured_source_binding": binding,
                "chunk_batch_ref_scoped": ":annual_audit:heavy:" in str(result.get("chunk_batch_ref") or ""),
                "final_report_ref_scoped": ":annual_audit:heavy:" in report_ref,
                "trace_item_count": len(trace_items),
                "evidence_count": evidence_count,
                "citation_coverage": result.get("citation_coverage"),
                "report_version": (artifacts.get("report") or {}).get("version"),
                "workpaper_count": len(artifacts.get("workpapers") or []),
                "artifact_status": artifacts.get("status"),
                "published_artifact_count": len(artifacts.get("artifacts") or []),
                "task_result": artifacts.get("tasks"),
            },
            ensure_ascii=False,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
