"""Content, provenance, filename, and container quality gates."""

from __future__ import annotations

import hashlib
import json
import posixpath
import re
import unicodedata
import zipfile
from io import BytesIO
from pathlib import PurePath
from typing import Any, Iterable, Literal

from defusedxml import ElementTree as SafeElementTree
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pydantic import BaseModel, ConfigDict, Field
from pypdf import PdfReader

from .content_schemas import (
    AttachmentContractError,
    BindingManifest,
    DocumentPayload,
    FactRefSegment,
    RenderResult,
    ResolvedDocumentPayload,
    ensure_payload_matches_manifest,
)
from .context_service import canonical_json_bytes


EXPECTED_CONTENT_TYPES = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".md": "text/markdown",
    ".pdf": "application/pdf",
}
_INTERNAL_PATTERNS = {
    "citation_wire_id": re.compile(r"\[\[\s*cite\s*:", re.IGNORECASE),
    "evidence_id": re.compile(r"\b(?:source_)?evidence[_-]?id\b", re.IGNORECASE),
    "chunk_id": re.compile(r"\b(?:source_)?chunk[_-]?id\b", re.IGNORECASE),
    "claim_id": re.compile(r"\bclaim[_-]?id\b", re.IGNORECASE),
    "object_ref": re.compile(r"\bminio://", re.IGNORECASE),
}
_PLACEHOLDER_PATTERNS = {
    "jinja_expression": re.compile(r"{{[\s\S]*?}}"),
    "jinja_statement": re.compile(r"{[%#][\s\S]*?[%#]}"),
    "dollar_placeholder": re.compile(r"\$\{[^{}]+}"),
    "angle_placeholder": re.compile(r"<<[^<>]+>>"),
    "bracket_placeholder": re.compile(r"\[\[[^\[\]]+]]"),
}
_INVALID_FILENAME_CHARS = re.compile(r"[<>:\"/\\|?*\x00-\x1f]")
_EXTERNAL_HYPERLINK_FORMULA = re.compile(
    r'^\s*=\s*(?:_xlfn\.)?HYPERLINK\s*\(\s*"(?:https?|ftp|file|mailto):',
    re.IGNORECASE,
)
_WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DOCX_COMMENT_TAGS = {
    "comment",
    "commentRangeStart",
    "commentRangeEnd",
    "commentReference",
}
_XLSX_PAGE_SETUP_ATTRIBUTES = (
    "orientation",
    "paperSize",
    "scale",
    "fitToHeight",
    "fitToWidth",
    "firstPageNumber",
    "useFirstPageNumber",
    "pageOrder",
    "blackAndWhite",
    "draft",
    "horizontalDpi",
    "verticalDpi",
)
_WINDOWS_RESERVED = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}


class QualityIssue(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    code: str
    severity: Literal["error", "warning"] = "error"
    message: str
    location: str = ""


class QualityReport(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    passed: bool
    file_name: str
    extension: str
    artifact_sha256: str
    extracted_text_sha256: str
    checks: list[str] = Field(default_factory=list)
    issues: list[QualityIssue] = Field(default_factory=list)

    @property
    def report_sha256(self) -> str:
        return hashlib.sha256(canonical_json_bytes(self.model_dump(mode="json"))).hexdigest()


class AttachmentQualityError(AttachmentContractError):
    def __init__(self, report: QualityReport) -> None:
        self.report = report
        messages = "; ".join(issue.message for issue in report.issues if issue.severity == "error")
        super().__init__(messages or "attachment quality gate failed")


def validate_payload_contract(
    payload: DocumentPayload | ResolvedDocumentPayload,
    manifest: BindingManifest,
) -> list[QualityIssue]:
    """Validate slot coverage and evidence before a renderer receives content."""

    issues: list[QualityIssue] = []
    try:
        ensure_payload_matches_manifest(payload, manifest)
    except AttachmentContractError as exc:
        return [QualityIssue(code="PAYLOAD_CONTRACT", message=str(exc))]

    if isinstance(payload, ResolvedDocumentPayload):
        for binding in manifest.slots:
            slot = payload.slot_map.get(binding.slot_id)
            if slot is None:
                continue
            if binding.required and _is_empty_slot_value(slot.value):
                issues.append(
                    QualityIssue(
                        code="REQUIRED_SLOT_EMPTY",
                        message=f"required slot resolved to an empty value: {slot.slot_id}",
                        location=slot.slot_id,
                    )
                )
            if _binding_requires_fact_refs(binding) and not slot.fact_refs:
                issues.append(
                    QualityIssue(
                        code="FACT_REF_REQUIRED",
                        message=f"slot requires an authoritative fact_ref: {slot.slot_id}",
                        location=slot.slot_id,
                    )
                )
        return issues

    for binding in manifest.slots:
        slot = payload.slot_map.get(binding.slot_id)
        if slot is None:
            continue
        require_fact_refs = _binding_requires_fact_refs(binding)
        if slot.kind == "scalar":
            has_fact = any(isinstance(segment, FactRefSegment) for segment in slot.segments)
            if require_fact_refs and not has_fact:
                issues.append(
                    QualityIssue(
                        code="FACT_REF_REQUIRED",
                        message=f"slot requires an authoritative fact_ref: {slot.slot_id}",
                        location=slot.slot_id,
                    )
                )
        elif slot.kind == "narrative_blocks":
            for block in slot.blocks:
                has_fact = any(isinstance(segment, FactRefSegment) for segment in block.segments)
                if block.requires_evidence and not has_fact and not block.evidence_refs:
                    issues.append(
                        QualityIssue(
                            code="NARRATIVE_EVIDENCE_MISSING",
                            message=f"narrative block has no fact or evidence: {block.block_id}",
                            location=f"{slot.slot_id}.{block.block_id}",
                        )
                    )
                if require_fact_refs and not has_fact:
                    issues.append(
                        QualityIssue(
                            code="FACT_REF_REQUIRED",
                            message=f"narrative block requires a fact_ref: {block.block_id}",
                            location=f"{slot.slot_id}.{block.block_id}",
                        )
                    )
        else:
            if slot.source_fact_ref:
                continue
            for row_index, row in enumerate(slot.rows):
                has_fact = any(
                    isinstance(segment, FactRefSegment)
                    for cell in row.cells.values()
                    for segment in cell.segments
                ) or bool(row.fact_refs)
                if require_fact_refs and not has_fact:
                    issues.append(
                        QualityIssue(
                            code="FACT_REF_REQUIRED",
                            message=f"table row requires a fact_ref: {slot.slot_id}[{row_index}]",
                            location=f"{slot.slot_id}[{row_index}]",
                        )
                    )
    return issues


def validate_rendered_artifact(
    result: RenderResult,
    *,
    file_name: str,
    manifest: BindingManifest,
    template_bytes: bytes | None = None,
    payload: ResolvedDocumentPayload | None = None,
) -> QualityReport:
    """Run deterministic publication gates over one rendered artifact."""

    issues: list[QualityIssue] = []
    checks = [
        "filename",
        "extension_content_type_signature",
        "format_reopen",
        "forbidden_output_patterns",
        "unresolved_placeholders",
    ]
    extension = result.extension.lower()
    safe_name = sanitize_filename(file_name)
    if safe_name != file_name:
        issues.append(
            QualityIssue(
                code="FILENAME_UNSAFE",
                message="artifact filename is not in canonical safe form",
                location=file_name,
            )
        )
    if not file_name.lower().endswith(extension):
        issues.append(
            QualityIssue(
                code="EXTENSION_MISMATCH",
                message=f"filename extension does not match renderer output: {extension}",
            )
        )
    expected_type = EXPECTED_CONTENT_TYPES.get(extension, "")
    if not expected_type or not result.content_type.lower().startswith(expected_type.lower()):
        issues.append(
            QualityIssue(
                code="CONTENT_TYPE_MISMATCH",
                message=f"content type does not match {extension}",
            )
        )
    if (
        manifest.source_template_sha256
        and result.source_template_sha256 != manifest.source_template_sha256
    ):
        issues.append(
            QualityIssue(
                code="TEMPLATE_SHA_MISMATCH",
                message="renderer source SHA does not match the binding manifest",
            )
        )

    text, format_issues = _extract_and_reopen(result.data, extension)
    issues.extend(format_issues)
    if extension in {".docx", ".xlsx"}:
        checks.append("ooxml_package_safety")
        issues.extend(_scan_ooxml_package_safety(result.data, extension))
    issues.extend(_scan_forbidden_text(text, manifest.forbidden_output_patterns))
    issues.extend(_scan_placeholders(text))
    if payload is not None:
        checks.append("resolved_payload")
        issues.extend(validate_payload_contract(payload, manifest))
        payload_text = json.dumps(payload.model_dump(mode="json"), ensure_ascii=False, default=str)
        issues.extend(_scan_forbidden_text(payload_text, manifest.forbidden_output_patterns))
        issues.extend(_scan_placeholders(payload_text))

    if (
        template_bytes is not None
        and manifest.allowed_modified_parts
        and extension in {".docx", ".xlsx"}
    ):
        checks.append("declared_package_parts")
        changed = set(_changed_zip_parts(template_bytes, result.data))
        unexpected = sorted(changed - set(manifest.allowed_modified_parts))
        if unexpected:
            issues.append(
                QualityIssue(
                    code="UNDECLARED_PART_CHANGE",
                    message=f"renderer modified undeclared package parts: {', '.join(unexpected)}",
                )
            )

    if template_bytes is not None and extension in {".docx", ".xlsx"}:
        checks.append("template_structure")
        issues.extend(_validate_template_structure(template_bytes, result.data, extension))

    issues = _deduplicate_issues(issues)
    passed = not any(issue.severity == "error" for issue in issues)
    return QualityReport(
        passed=passed,
        file_name=file_name,
        extension=extension,
        artifact_sha256=hashlib.sha256(result.data).hexdigest(),
        extracted_text_sha256=hashlib.sha256(text.encode("utf-8")).hexdigest(),
        checks=checks,
        issues=issues,
    )


def require_quality(report: QualityReport) -> QualityReport:
    if not report.passed:
        raise AttachmentQualityError(report)
    return report


def sanitize_output_filename(
    project_name: str,
    display_name: str,
    extension: str,
    *,
    sequence: int | None = None,
) -> str:
    normalized_extension = extension.lower()
    if normalized_extension not in EXPECTED_CONTENT_TYPES:
        raise ValueError(f"unsupported attachment extension: {extension}")
    suffix = str(sequence) if sequence is not None else ""
    return sanitize_filename(f"{project_name}{display_name}{suffix}{normalized_extension}")


def sanitize_filename(file_name: str, *, max_length: int = 180) -> str:
    normalized = unicodedata.normalize("NFKC", str(file_name or ""))
    normalized = normalized.replace("\\", "/").rsplit("/", 1)[-1]
    normalized = _INVALID_FILENAME_CHARS.sub("_", normalized)
    normalized = re.sub(r"_+", "_", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip(" .")
    if not normalized:
        normalized = "attachment"
    path = PurePath(normalized)
    stem = path.stem.strip(" .") or "attachment"
    suffix = path.suffix
    if stem.upper() in _WINDOWS_RESERVED:
        stem = f"_{stem}"
    room = max(1, max_length - len(suffix))
    stem = stem[:room].rstrip(" .") or "attachment"
    return f"{stem}{suffix.lower()}"


def _extract_and_reopen(data: bytes, extension: str) -> tuple[str, list[QualityIssue]]:
    issues: list[QualityIssue] = []
    try:
        if extension == ".docx":
            return _extract_docx(data), issues
        if extension == ".xlsx":
            return _extract_xlsx(data, issues), issues
        if extension == ".md":
            return data.decode("utf-8"), issues
        if extension == ".pdf":
            reader = PdfReader(BytesIO(data), strict=True)
            if reader.is_encrypted:
                raise ValueError("rendered PDF is encrypted")
            return "\n".join(page.extract_text() or "" for page in reader.pages), issues
        raise ValueError(f"unsupported extension: {extension}")
    except Exception as exc:
        issues.append(
            QualityIssue(
                code="FORMAT_REOPEN_FAILED",
                message=f"rendered {extension} cannot be reopened: {type(exc).__name__}",
            )
        )
        return "", issues


def _extract_docx(data: bytes) -> str:
    with zipfile.ZipFile(BytesIO(data), "r") as archive:
        names = set(archive.namelist())
        required = {"[Content_Types].xml", "word/document.xml"}
        if not required.issubset(names):
            raise ValueError("DOCX package is incomplete")
        parts = [
            name
            for name in names
            if name.startswith("word/") and name.endswith(".xml")
        ]
        values: list[str] = []
        for name in sorted(parts):
            root = SafeElementTree.fromstring(archive.read(name))
            for element in root.iter():
                if element.tag.rsplit("}", 1)[-1] in {"t", "instrText"} and element.text:
                    values.append(element.text)
        return "\n".join(values)


def _extract_xlsx(data: bytes, issues: list[QualityIssue]) -> str:
    workbook = load_workbook(BytesIO(data), data_only=False, read_only=False, keep_links=False)
    values: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            values.append(worksheet.title)
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        issues.append(
                            QualityIssue(
                                code="AUTHOR_COMMENT_PRESENT",
                                message="rendered XLSX contains an author comment",
                                location=f"{worksheet.title}!{cell.coordinate}",
                            )
                        )
                    if isinstance(cell.value, str) and _EXTERNAL_HYPERLINK_FORMULA.search(
                        cell.value
                    ):
                        issues.append(
                            QualityIssue(
                                code="EXTERNAL_LINK_PRESENT",
                                message="rendered XLSX contains an external hyperlink formula",
                                location=f"{worksheet.title}!{cell.coordinate}",
                            )
                        )
                    if cell.value is not None:
                        values.append(str(cell.value))
    finally:
        workbook.close()
    return "\n".join(values)


def _scan_ooxml_package_safety(data: bytes, extension: str) -> list[QualityIssue]:
    issues: list[QualityIssue] = []
    try:
        with zipfile.ZipFile(BytesIO(data), "r") as archive:
            names = set(archive.namelist())
            for name in sorted(item for item in names if item.endswith(".rels")):
                try:
                    root = SafeElementTree.fromstring(archive.read(name))
                except Exception:
                    issues.append(
                        QualityIssue(
                            code="OOXML_RELATIONSHIP_INVALID",
                            message="rendered OOXML contains an invalid relationship part",
                            location=name,
                        )
                    )
                    continue
                for relationship in root.iter():
                    if _local_name(relationship.tag) != "Relationship":
                        continue
                    if relationship.attrib.get("TargetMode", "").lower() != "external":
                        continue
                    issues.append(
                        QualityIssue(
                            code="EXTERNAL_LINK_PRESENT",
                            message="rendered OOXML contains an external relationship",
                            location=name,
                        )
                    )

            if extension == ".docx":
                for name in sorted(
                    item
                    for item in names
                    if item.startswith("word/") and item.endswith(".xml")
                ):
                    try:
                        root = SafeElementTree.fromstring(archive.read(name))
                    except Exception:
                        continue
                    if any(
                        _local_name(element.tag) in _DOCX_COMMENT_TAGS
                        for element in root.iter()
                    ):
                        issues.append(
                            QualityIssue(
                                code="AUTHOR_COMMENT_PRESENT",
                                message="rendered DOCX contains an author comment",
                                location=name,
                            )
                        )

            if extension == ".xlsx":
                for name in sorted(
                    item
                    for item in names
                    if item.startswith("xl/externalLinks/") and not item.endswith("/")
                ):
                    issues.append(
                        QualityIssue(
                            code="EXTERNAL_LINK_PRESENT",
                            message="rendered XLSX contains an external workbook link",
                            location=name,
                        )
                    )
                for name in sorted(
                    item
                    for item in names
                    if item.startswith("xl/threadedComments/") and item.endswith(".xml")
                ):
                    issues.append(
                        QualityIssue(
                            code="AUTHOR_COMMENT_PRESENT",
                            message="rendered XLSX contains a threaded author comment",
                            location=name,
                        )
                    )
    except zipfile.BadZipFile:
        # The format-reopen gate reports the malformed container.
        return []
    return _deduplicate_issues(issues)


def _validate_template_structure(
    template_bytes: bytes,
    rendered_bytes: bytes,
    extension: str,
) -> list[QualityIssue]:
    try:
        if extension == ".docx":
            return _compare_docx_structure(template_bytes, rendered_bytes)
        if extension == ".xlsx":
            return _compare_xlsx_structure(template_bytes, rendered_bytes)
    except Exception as exc:
        return [
            QualityIssue(
                code="TEMPLATE_STRUCTURE_CHECK_FAILED",
                message=(
                    f"{extension} template structure could not be compared: "
                    f"{type(exc).__name__}"
                ),
            )
        ]
    return []


def _compare_docx_structure(template_bytes: bytes, rendered_bytes: bytes) -> list[QualityIssue]:
    template = _docx_structure_snapshot(template_bytes)
    rendered = _docx_structure_snapshot(rendered_bytes)
    issues: list[QualityIssue] = []
    if template["sections"] != rendered["sections"]:
        issues.append(
            QualityIssue(
                code="DOCX_SECTION_STRUCTURE_CHANGED",
                message=(
                    "rendered DOCX changed section, page, or header/footer reference structure"
                ),
            )
        )
    if template["header_footer_parts"] != rendered["header_footer_parts"]:
        issues.append(
            QualityIssue(
                code="DOCX_HEADER_FOOTER_STRUCTURE_CHANGED",
                message="rendered DOCX changed referenced header/footer structural markup",
            )
        )

    rendered_styles = rendered["styles"]
    missing_or_changed = sorted(
        style_id
        for style_id, signature in template["styles"].items()
        if rendered_styles.get(style_id) != signature
    )
    if (
        missing_or_changed
        or template["style_defaults"] != rendered["style_defaults"]
    ):
        location = ", ".join(missing_or_changed[:10])
        issues.append(
            QualityIssue(
                code="DOCX_STYLE_STRUCTURE_CHANGED",
                message="rendered DOCX removed or changed template style structure",
                location=location,
            )
        )
    return issues


def _docx_structure_snapshot(data: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(BytesIO(data), "r") as archive:
        names = set(archive.namelist())
        document = SafeElementTree.fromstring(archive.read("word/document.xml"))
        relationships = _docx_document_relationships(archive, names)
        sections: list[Any] = []
        header_footer_parts: list[Any] = []
        for section_index, section in enumerate(document.iter(_word_q("sectPr"))):
            reference_health: list[tuple[str, str, bool, bool]] = []
            for reference_index, reference in enumerate(list(section)):
                kind = _local_name(reference.tag)
                if kind not in {"headerReference", "footerReference"}:
                    continue
                rel_id = reference.attrib.get(f"{{{_REL_NS}}}id", "")
                relationship = relationships.get(rel_id)
                target = relationship[0] if relationship is not None else ""
                rel_type = relationship[1] if relationship is not None else ""
                normalized_target = _word_relationship_target(target)
                target_exists = normalized_target in names
                part_signature = None
                if target_exists:
                    part_root = SafeElementTree.fromstring(archive.read(normalized_target))
                    part_signature = _canonical_word_structure(
                        part_root,
                        ignore_text_attributes=True,
                    )
                reference_health.append(
                    (
                        kind,
                        reference.attrib.get(_word_q("type"), "default"),
                        rel_type.rsplit("/", 1)[-1] == kind.removesuffix("Reference"),
                        target_exists,
                    )
                )
                header_footer_parts.append(
                    (
                        section_index,
                        reference_index,
                        kind,
                        reference.attrib.get(_word_q("type"), "default"),
                        part_signature,
                    )
                )
            sections.append(
                (
                    _canonical_word_structure(section),
                    tuple(reference_health),
                )
            )

        styles: dict[str, Any] = {}
        style_defaults: list[Any] = []
        if "word/styles.xml" in names:
            styles_root = SafeElementTree.fromstring(archive.read("word/styles.xml"))
            for child in list(styles_root):
                kind = _local_name(child.tag)
                if kind == "style":
                    style_id = child.attrib.get(_word_q("styleId"), "")
                    styles[style_id] = _canonical_word_structure(child)
                elif kind in {"docDefaults", "latentStyles"}:
                    style_defaults.append(_canonical_word_structure(child))
        return {
            "sections": tuple(sections),
            "header_footer_parts": tuple(header_footer_parts),
            "styles": styles,
            "style_defaults": tuple(style_defaults),
        }


def _docx_document_relationships(
    archive: zipfile.ZipFile,
    names: set[str],
) -> dict[str, tuple[str, str]]:
    relationship_part = "word/_rels/document.xml.rels"
    if relationship_part not in names:
        return {}
    root = SafeElementTree.fromstring(archive.read(relationship_part))
    return {
        relationship.attrib.get("Id", ""): (
            relationship.attrib.get("Target", ""),
            relationship.attrib.get("Type", ""),
        )
        for relationship in root.iter()
        if _local_name(relationship.tag) == "Relationship"
        and relationship.attrib.get("TargetMode", "").lower() != "external"
    }


def _word_relationship_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("word", target))


def _canonical_word_structure(
    element: Any,
    *,
    ignore_text_attributes: bool = False,
) -> tuple[Any, ...]:
    kind = _local_name(element.tag)
    attributes = tuple(
        sorted(
            (_local_name(name), value)
            for name, value in element.attrib.items()
            if not (ignore_text_attributes and kind in {"t", "instrText", "delText"})
            if not _local_name(name).startswith("rsid")
            and not (
                kind in {"headerReference", "footerReference"}
                and _local_name(name) == "id"
            )
        )
    )
    return (
        kind,
        attributes,
        tuple(
            _canonical_word_structure(
                child,
                ignore_text_attributes=ignore_text_attributes,
            )
            for child in list(element)
        ),
    )


def _compare_xlsx_structure(template_bytes: bytes, rendered_bytes: bytes) -> list[QualityIssue]:
    template = _xlsx_structure_snapshot(template_bytes)
    rendered = _xlsx_structure_snapshot(rendered_bytes)
    issues: list[QualityIssue] = []
    if template["sheetnames"] != rendered["sheetnames"]:
        issues.append(
            QualityIssue(
                code="XLSX_SHEET_STRUCTURE_CHANGED",
                message="rendered XLSX changed template worksheet names or order",
            )
        )

    for sheet_name in template["sheetnames"]:
        before = template["worksheets"].get(sheet_name)
        after = rendered["worksheets"].get(sheet_name)
        if before is None or after is None:
            continue
        facets = (
            ("merges", "XLSX_MERGE_STRUCTURE_CHANGED", "merged-cell structure"),
            ("print_settings", "XLSX_PRINT_SETTINGS_CHANGED", "print area or print titles"),
            ("page_setup", "XLSX_PAGE_SETUP_CHANGED", "page setup"),
            ("header_footer", "XLSX_HEADER_FOOTER_CHANGED", "header/footer structure"),
            ("freeze_panes", "XLSX_FREEZE_PANES_CHANGED", "freeze panes"),
            ("tables", "XLSX_TABLE_STRUCTURE_CHANGED", "Excel Table structure"),
            (
                "defined_names",
                "XLSX_DEFINED_NAME_STRUCTURE_CHANGED",
                "worksheet defined-name structure",
            ),
        )
        for key, code, label in facets:
            if before[key] != after[key]:
                issues.append(
                    QualityIssue(
                        code=code,
                        message=f"rendered XLSX changed template {label}",
                        location=sheet_name,
                    )
                )
    if template["defined_names"] != rendered["defined_names"]:
        issues.append(
            QualityIssue(
                code="XLSX_DEFINED_NAME_STRUCTURE_CHANGED",
                message="rendered XLSX changed workbook defined-name structure",
            )
        )
    return issues


def _xlsx_structure_snapshot(data: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(data), data_only=False, read_only=False, keep_links=False)
    try:
        worksheets: dict[str, dict[str, Any]] = {}
        for worksheet in workbook.worksheets:
            freeze_panes = worksheet.freeze_panes
            if hasattr(freeze_panes, "coordinate"):
                freeze_panes = freeze_panes.coordinate
            worksheets[worksheet.title] = {
                "merges": tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
                "print_settings": (
                    str(worksheet.print_area or ""),
                    str(worksheet.print_title_rows or ""),
                    str(worksheet.print_title_cols or ""),
                ),
                "page_setup": (
                    tuple(
                        getattr(worksheet.page_setup, attribute)
                        for attribute in _XLSX_PAGE_SETUP_ATTRIBUTES
                    ),
                    worksheet.sheet_properties.pageSetUpPr.fitToPage,
                    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks,
                ),
                "header_footer": _xlsx_header_footer_signature(worksheet),
                "freeze_panes": str(freeze_panes or ""),
                "tables": _xlsx_table_signatures(worksheet),
                "defined_names": _xlsx_defined_name_signatures(worksheet.defined_names),
            }
        return {
            "sheetnames": tuple(workbook.sheetnames),
            "worksheets": worksheets,
            "defined_names": _xlsx_defined_name_signatures(workbook.defined_names),
        }
    finally:
        workbook.close()


def _xlsx_header_footer_signature(worksheet: Any) -> tuple[Any, ...]:
    items: list[Any] = [
        worksheet.HeaderFooter.differentOddEven,
        worksheet.HeaderFooter.differentFirst,
        worksheet.HeaderFooter.scaleWithDoc,
        worksheet.HeaderFooter.alignWithMargins,
    ]
    for item_name in (
        "oddHeader",
        "oddFooter",
        "evenHeader",
        "evenFooter",
        "firstHeader",
        "firstFooter",
    ):
        item = getattr(worksheet, item_name)
        items.append(
            tuple(
                tuple(
                    getattr(getattr(item, position), attribute)
                    for attribute in ("text", "font", "size", "color")
                )
                for position in ("left", "center", "right")
            )
        )
    return tuple(items)


def _xlsx_table_signatures(worksheet: Any) -> tuple[Any, ...]:
    signatures: list[Any] = []
    for table in worksheet.tables.values():
        min_column, min_row, max_column, _ = range_boundaries(table.ref)
        style = table.tableStyleInfo
        style_signature = None
        if style is not None:
            style_signature = (
                style.name,
                style.showFirstColumn,
                style.showLastColumn,
                style.showRowStripes,
                style.showColumnStripes,
            )
        column_signatures = tuple(
            (
                column.id,
                column.name,
                column.totalsRowFunction,
                column.totalsRowLabel,
                _xlsx_formula_text(column.calculatedColumnFormula),
                _xlsx_formula_text(column.totalsRowFormula),
            )
            for column in table.tableColumns
        )
        signatures.append(
            (
                table.name,
                table.displayName,
                min_column,
                min_row,
                max_column,
                table.tableType,
                table.headerRowCount,
                table.totalsRowCount,
                table.totalsRowShown,
                column_signatures,
                style_signature,
            )
        )
    return tuple(sorted(signatures, key=repr))


def _xlsx_formula_text(formula: Any) -> str:
    if formula is None:
        return ""
    return str(getattr(formula, "text", formula) or "")


def _xlsx_defined_name_signatures(defined_names: Any) -> tuple[Any, ...]:
    signatures = [
        (
            name,
            defined_name.attr_text or "",
            defined_name.localSheetId,
            defined_name.hidden,
            defined_name.function,
            defined_name.vbProcedure,
            defined_name.xlm,
            defined_name.workbookParameter,
        )
        for name, defined_name in defined_names.items()
    ]
    return tuple(sorted(signatures, key=repr))


def _word_q(local_name: str) -> str:
    return f"{{{_WORD_NS}}}{local_name}"


def _local_name(value: str) -> str:
    return value.rsplit("}", 1)[-1]


def _scan_forbidden_text(text: str, custom_patterns: Iterable[str]) -> list[QualityIssue]:
    issues: list[QualityIssue] = []
    for code, pattern in _INTERNAL_PATTERNS.items():
        if pattern.search(text):
            issues.append(
                QualityIssue(
                    code="FORBIDDEN_INTERNAL_ID",
                    message=f"client artifact contains internal marker: {code}",
                )
            )
    lowered = text.lower()
    for pattern in custom_patterns:
        normalized = str(pattern or "").strip()
        if normalized and normalized.lower() in lowered:
            issues.append(
                QualityIssue(
                    code="FORBIDDEN_CUSTOM_PATTERN",
                    message=f"client artifact contains forbidden pattern: {normalized}",
                )
            )
    return _deduplicate_issues(issues)


def _scan_placeholders(text: str) -> list[QualityIssue]:
    return [
        QualityIssue(
            code="UNRESOLVED_PLACEHOLDER",
            message=f"client artifact contains unresolved placeholder: {name}",
        )
        for name, pattern in _PLACEHOLDER_PATTERNS.items()
        if pattern.search(text)
    ]


def _binding_requires_fact_refs(binding) -> bool:
    if "require_fact_refs" in binding.options:
        return bool(binding.options["require_fact_refs"])
    source = str(binding.source or "").removeprefix("document.")
    return source.startswith(("entity.", "entity_facts.", "facts."))


def _is_empty_slot_value(value: Any) -> bool:
    return value is None or value == "" or value == [] or value == {}


def _changed_zip_parts(before: bytes, after: bytes) -> list[str]:
    before_hashes = _zip_hashes(before)
    after_hashes = _zip_hashes(after)
    return sorted(
        name
        for name in set(before_hashes) | set(after_hashes)
        if before_hashes.get(name) != after_hashes.get(name)
    )


def _zip_hashes(data: bytes) -> dict[str, str]:
    with zipfile.ZipFile(BytesIO(data), "r") as archive:
        return {
            name: hashlib.sha256(archive.read(name)).hexdigest()
            for name in archive.namelist()
        }


def _deduplicate_issues(issues: list[QualityIssue]) -> list[QualityIssue]:
    result: list[QualityIssue] = []
    seen: set[tuple[str, str, str]] = set()
    for issue in issues:
        key = (issue.code, issue.message, issue.location)
        if key not in seen:
            result.append(issue)
            seen.add(key)
    return result


validate_artifact = validate_rendered_artifact


__all__ = [
    "AttachmentQualityError",
    "EXPECTED_CONTENT_TYPES",
    "QualityIssue",
    "QualityReport",
    "require_quality",
    "sanitize_filename",
    "sanitize_output_filename",
    "validate_artifact",
    "validate_payload_contract",
    "validate_rendered_artifact",
]
