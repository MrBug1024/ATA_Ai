"""Deterministic XLSX rendering for explicit cells, named ranges, and tables."""

from __future__ import annotations

import copy
import hashlib
import json
import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries

from ..content_schemas import (
    AttachmentContractError,
    AttachmentRenderError,
    BindingManifest,
    RenderResult,
    ResolvedDocumentPayload,
    SlotBinding,
    ensure_payload_matches_manifest,
    verify_source_template,
)


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def render_xlsx(
    template_bytes: bytes,
    payload: ResolvedDocumentPayload,
    manifest: BindingManifest,
) -> RenderResult:
    """Write only the ranges named by the confirmed XLSX binding manifest."""

    ensure_payload_matches_manifest(payload, manifest)
    source_sha = verify_source_template(template_bytes, manifest)
    _assert_xlsx_container(template_bytes)
    try:
        workbook = load_workbook(BytesIO(template_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise AttachmentContractError("XLSX template cannot be opened") from exc

    for binding in manifest.slots:
        slot = payload.slot_map.get(binding.slot_id)
        if slot is None:
            continue
        format_name, target_kind, locator = binding.target.split(":", 2)
        if format_name != "xlsx":
            raise AttachmentContractError(f"non-XLSX target in XLSX manifest: {binding.target}")
        if target_kind == "cell":
            _write_cell_target(workbook, locator, slot.kind, slot.value, binding)
        elif target_kind == "named-range":
            _write_named_range(workbook, locator, slot.kind, slot.value, binding)
        elif target_kind == "range":
            _write_explicit_range(workbook, locator, slot.kind, slot.value, binding)
        elif target_kind == "table":
            _write_table(workbook, locator, slot.kind, slot.value, binding)
        else:
            raise AttachmentContractError(f"unsupported XLSX binding target: {binding.target}")

    output = BytesIO()
    try:
        workbook.save(output)
        rendered = output.getvalue()
        load_workbook(BytesIO(rendered), data_only=False, read_only=True, keep_links=False).close()
    except Exception as exc:
        raise AttachmentRenderError("rendered XLSX cannot be reopened") from exc
    finally:
        workbook.close()

    return RenderResult(
        data=rendered,
        extension=".xlsx",
        content_type=XLSX_CONTENT_TYPE,
        source_template_sha256=source_sha,
        modified_parts=tuple(_changed_zip_parts(template_bytes, rendered)),
    )


def escape_formula_text(value: Any) -> Any:
    """Prevent user-controlled strings from becoming Excel formulas."""

    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _write_cell_target(workbook, locator: str, kind: str, value: Any, binding: SlotBinding) -> None:
    worksheet, coordinate = _resolve_cell(workbook, locator, binding)
    if kind == "table_rows":
        raise AttachmentContractError("table_rows cannot be written to a single XLSX cell")
    worksheet[coordinate] = _excel_value(_scalar_value(kind, value))


def _write_named_range(workbook, name: str, kind: str, value: Any, binding: SlotBinding) -> None:
    defined_name = workbook.defined_names.get(name)
    if defined_name is None:
        raise AttachmentContractError(f"XLSX named range was not found: {name}")
    try:
        destinations = list(defined_name.destinations)
    except Exception as exc:
        raise AttachmentContractError(f"XLSX defined name is not a cell range: {name}") from exc
    if len(destinations) != 1 and not bool(binding.options.get("allow_multiple", False)):
        raise AttachmentContractError(
            f"XLSX named range must resolve to exactly one destination: {name}"
        )
    for sheet_name, cell_range in destinations:
        if sheet_name not in workbook.sheetnames:
            raise AttachmentContractError(
                f"XLSX named range refers to a missing worksheet: {sheet_name}"
            )
        worksheet = workbook[sheet_name]
        _write_fixed_range(worksheet, cell_range, kind, value, binding)


def _write_explicit_range(
    workbook,
    locator: str,
    kind: str,
    value: Any,
    binding: SlotBinding,
) -> None:
    if "!" in locator:
        sheet_name, cell_range = locator.rsplit("!", 1)
        sheet_name = sheet_name.strip("'")
    else:
        sheet_name = str(binding.options.get("sheet") or "")
        cell_range = locator
    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise AttachmentContractError(
            f"XLSX range target has an invalid worksheet: {locator}"
        )
    try:
        range_boundaries(cell_range)
    except ValueError as exc:
        raise AttachmentContractError(f"XLSX range is invalid: {locator}") from exc
    _write_fixed_range(workbook[sheet_name], cell_range, kind, value, binding)


def _write_fixed_range(worksheet, cell_range: str, kind: str, value: Any, binding: SlotBinding) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if kind != "table_rows":
        worksheet.cell(min_row, min_col).value = _excel_value(_scalar_value(kind, value))
        return
    rows = list(value or [])
    fields = _column_fields(binding, [], rows)
    available_rows = max_row - min_row + 1
    if len(rows) > available_rows:
        if binding.overflow_policy == "truncate":
            rows = rows[:available_rows]
        else:
            raise AttachmentContractError(
                f"table payload exceeds fixed named range {cell_range}"
            )
    if len(fields) > max_col - min_col + 1:
        raise AttachmentContractError(f"table payload has too many columns for {cell_range}")
    for row_offset in range(available_rows):
        source = rows[row_offset] if row_offset < len(rows) else {}
        for column_offset, field in enumerate(fields):
            worksheet.cell(min_row + row_offset, min_col + column_offset).value = _excel_value(
                source.get(field, "")
            )


def _write_table(workbook, table_name: str, kind: str, value: Any, binding: SlotBinding) -> None:
    if kind != "table_rows":
        raise AttachmentContractError("XLSX table targets require table_rows payloads")
    matches = [
        (worksheet, worksheet.tables[table_name])
        for worksheet in workbook.worksheets
        if table_name in worksheet.tables
    ]
    if len(matches) != 1:
        raise AttachmentContractError(
            f"XLSX table must exist exactly once: {table_name}"
        )
    worksheet, table = matches[0]
    min_col, header_row, max_col, old_last_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(header_row, column).value or "").strip()
        for column in range(min_col, max_col + 1)
    ]
    rows = list(value or [])
    fields = _column_fields(binding, headers, rows)
    if len(fields) != len(headers):
        raise AttachmentContractError(
            f"XLSX table {table_name} column contract does not match template headers"
        )

    prototype_row = int(binding.options.get("prototype_row") or min(header_row + 1, old_last_row))
    if prototype_row <= header_row:
        raise AttachmentContractError(f"XLSX table {table_name} requires a prototype data row")
    formula_columns = {
        str(item)
        for item in (binding.options.get("formula_columns") or [])
    }
    new_last_row = header_row + len(rows)
    _assert_table_extension_region_available(
        worksheet,
        table_name=table_name,
        min_col=min_col,
        max_col=max_col,
        old_last_row=old_last_row,
        new_last_row=new_last_row,
    )
    for row_index, row_value in enumerate(rows, start=header_row + 1):
        if row_index != prototype_row:
            _clone_row_style(worksheet, prototype_row, row_index, min_col, max_col)
        for column_offset, field in enumerate(fields):
            column = min_col + column_offset
            header = headers[column_offset]
            if (field in formula_columns or header in formula_columns) and field not in row_value:
                prototype = worksheet.cell(prototype_row, column)
                formula = prototype.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise AttachmentContractError(
                        f"declared formula column has no prototype formula: {header}"
                    )
                target = worksheet.cell(row_index, column)
                target.value = Translator(
                    formula,
                    origin=prototype.coordinate,
                ).translate_formula(target.coordinate)
            else:
                worksheet.cell(row_index, column).value = _excel_value(
                    row_value.get(field, "")
                )

    for row_index in range(max(header_row + 1, new_last_row + 1), old_last_row + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row_index, column).value = None

    if not rows:
        if binding.required:
            raise AttachmentContractError(f"required XLSX table {table_name} has no rows")
        new_last_row = header_row + 1
        for column in range(min_col, max_col + 1):
            worksheet.cell(new_last_row, column).value = None
    table.ref = (
        f"{get_column_letter(min_col)}{header_row}:"
        f"{get_column_letter(max_col)}{max(new_last_row, header_row + 1)}"
    )


def _assert_table_extension_region_available(
    worksheet,
    *,
    table_name: str,
    min_col: int,
    max_col: int,
    old_last_row: int,
    new_last_row: int,
) -> None:
    """Never grow a table across content owned by another template region."""

    if new_last_row <= old_last_row:
        return
    extension_min_row = old_last_row + 1
    for row in worksheet.iter_rows(
        min_row=extension_min_row,
        max_row=new_last_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if cell.value is not None or cell.comment is not None or cell.hyperlink is not None:
                raise AttachmentContractError(
                    f"XLSX table {table_name} cannot extend across existing content at "
                    f"{cell.coordinate}"
                )

    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.max_row >= extension_min_row
            and merged_range.min_row <= new_last_row
            and merged_range.max_col >= min_col
            and merged_range.min_col <= max_col
        ):
            raise AttachmentContractError(
                f"XLSX table {table_name} cannot extend across merged range "
                f"{merged_range.coord}"
            )

    for other_table in worksheet.tables.values():
        if other_table.name == table_name:
            continue
        other_min_col, other_min_row, other_max_col, other_max_row = range_boundaries(
            other_table.ref
        )
        if (
            other_max_row >= extension_min_row
            and other_min_row <= new_last_row
            and other_max_col >= min_col
            and other_min_col <= max_col
        ):
            raise AttachmentContractError(
                f"XLSX table {table_name} cannot extend across table {other_table.name}"
            )


def _column_fields(binding: SlotBinding, headers: list[str], rows: list[dict[str, Any]]) -> list[str]:
    column_map = binding.options.get("column_map")
    if isinstance(column_map, dict):
        if headers:
            missing_headers = [header for header in headers if header not in column_map]
            if missing_headers:
                raise AttachmentContractError(
                    f"XLSX column_map is missing headers: {', '.join(missing_headers)}"
                )
            return [str(column_map[header]) for header in headers]
        return [str(value) for value in column_map.values()]
    columns = binding.options.get("columns")
    if isinstance(columns, list) and all(isinstance(item, str) for item in columns):
        return list(columns)
    if headers and all(not rows or header in rows[0] for header in headers):
        return headers
    if rows:
        return list(rows[0].keys())
    return headers


def _clone_row_style(worksheet, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(min_col, max_col + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)


def _resolve_cell(workbook, locator: str, binding: SlotBinding):
    if "!" in locator:
        sheet_name, coordinate = locator.rsplit("!", 1)
        sheet_name = sheet_name.strip("'")
    else:
        sheet_name = str(binding.options.get("sheet") or "")
        coordinate = locator
    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise AttachmentContractError(f"XLSX cell target has an invalid worksheet: {locator}")
    try:
        range_boundaries(coordinate)
    except ValueError as exc:
        raise AttachmentContractError(f"XLSX cell coordinate is invalid: {locator}") from exc
    return workbook[sheet_name], coordinate


def _scalar_value(kind: str, value: Any) -> Any:
    if kind == "narrative_blocks":
        return "\n".join(str(block.get("text") or "") for block in value)
    return value


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, Decimal, date, datetime)):
        return value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return escape_formula_text(str(value))


def _assert_xlsx_container(data: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(data), "r") as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile as exc:
        raise AttachmentContractError("XLSX template is not a valid ZIP container") from exc
    required = {"[Content_Types].xml", "xl/workbook.xml"}
    missing = sorted(required - names)
    if missing:
        raise AttachmentContractError(
            f"XLSX template is missing required package parts: {', '.join(missing)}"
        )


def _changed_zip_parts(before: bytes, after: bytes) -> list[str]:
    before_hashes = _zip_hashes(before)
    after_hashes = _zip_hashes(after)
    return sorted(
        name
        for name in set(before_hashes) | set(after_hashes)
        if before_hashes.get(name) != after_hashes.get(name)
    )


def _zip_hashes(data: bytes) -> dict[str, str]:
    with zipfile.ZipFile(BytesIO(data), "r") as archive:
        return {
            name: hashlib.sha256(archive.read(name)).hexdigest()
            for name in archive.namelist()
        }


__all__ = ["XLSX_CONTENT_TYPE", "escape_formula_text", "render_xlsx"]
