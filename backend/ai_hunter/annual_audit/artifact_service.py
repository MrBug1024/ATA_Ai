"""Versioned annual-audit report/workpaper artifact rendering.

The renderer is intentionally deterministic.  LLM-generated prose is treated
as input text; the artifact metadata, version and source snapshot remain
authoritative and are persisted separately.
"""

from __future__ import annotations

import html
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from xml.sax.saxutils import escape

from ai_hunter.app.services.minio_service import get_minio_service
from ai_hunter.platform_core import ArtifactRef


REPORT_CONTENT_TYPES = {
    "md": "text/markdown; charset=utf-8",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

_CITATION_WIRE_PATTERN = re.compile(r"\[\[cite:([1-9]\d*)\]\]")


def _artifact_display_text(report_text: str) -> str:
    """Materialize explicit chat citation markers for non-interactive files."""

    return _CITATION_WIRE_PATTERN.sub(r"[\1]", report_text)


def _json_safe(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _flatten(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    if isinstance(value, dict):
        rows: list[tuple[str, str]] = []
        for key, item in value.items():
            rows.extend(_flatten(item, f"{prefix}.{key}" if prefix else str(key)))
        return rows
    if isinstance(value, list):
        return [(prefix, str(value))]
    return [(prefix, str(_json_safe(value) if value is not None else ""))]


def build_workpaper_xlsx(*, code: str, name: str, facts: dict[str, Any], version: int) -> bytes:
    """Build a portable workbook with facts, conclusion and source locators."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底稿"
    sheet.append(["工作底稿编号", code])
    sheet.append(["工作底稿名称", name])
    sheet.append(["工作底稿版本", version])
    sheet.append([])
    sheet.append(["字段", "值"])
    for cell in sheet[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for key, value in _flatten(facts):
        sheet.append([key, value])
    sheet.freeze_panes = "A6"
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 120
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_report_xlsx(*, report_text: str, snapshot: dict[str, Any], report_version: int) -> bytes:
    """Build a review workbook containing a draft, findings and evidence."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "审计报告草稿"
    report_sheet.append(["年度财务报表审计报告草稿版本", report_version])
    report_sheet.append(["模板版本", snapshot.get("report_template_version", "")])
    report_sheet.append([])
    for line in report_text.splitlines():
        report_sheet.append([line])
    report_sheet.column_dimensions["A"].width = 140
    report_sheet.freeze_panes = "A4"

    finding_sheet = workbook.create_sheet("审计发现")
    finding_sheet.append(["分析类型", "风险等级", "标题", "描述", "金额", "证据数量"])
    for cell in finding_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for analysis_type, result_key in (
        ("sales_receivables", "sales_receivables"),
        ("cash_and_bank", "cash_and_bank"),
    ):
        result = snapshot.get(result_key) or {}
        for finding in result.get("findings") or []:
            finding_sheet.append(
                [
                    analysis_type,
                    finding.get("risk_level", ""),
                    finding.get("title", ""),
                    finding.get("description", ""),
                    finding.get("amount", ""),
                    len(finding.get("evidence_refs") or []),
                ]
            )
    finding_sheet.column_dimensions["D"].width = 90

    readiness_sheet = workbook.create_sheet("资料完整性")
    readiness_sheet.append(["项目", "值"])
    for key, value in _flatten(snapshot.get("readiness") or {}):
        readiness_sheet.append([key, value])
    readiness_sheet.column_dimensions["A"].width = 45
    readiness_sheet.column_dimensions["B"].width = 100

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_docx(*, title: str, report_text: str) -> bytes:
    """Create a minimal valid DOCX without introducing a heavyweight runtime dependency."""

    paragraphs = [title, *report_text.splitlines()]
    body = "".join(
        f'<w:p><w:r><w:t xml:space="preserve">{escape(line or " ")}</w:t></w:r></w:p>'
        for line in paragraphs
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}<w:sectPr/></w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '</Relationships>'
    )
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("word/document.xml", document)
    return output.getvalue()


def publish_annual_artifacts(
    *,
    engagement_id: int,
    report_text: str,
    snapshot: dict[str, Any],
    report_version: int,
    workpapers: list[dict[str, Any]],
) -> dict[str, Any]:
    """Render and persist review-only draft artifacts to the annual bucket."""

    service = get_minio_service()
    report_name = f"annual-audit-draft-{engagement_id}-v{report_version}"
    artifact_report_text = _artifact_display_text(report_text)
    payloads: list[tuple[str, str, bytes, str, int]] = [
        (
            f"{report_name}.md",
            REPORT_CONTENT_TYPES["md"],
            artifact_report_text.encode("utf-8"),
            "annual_report_markdown",
            report_version,
        ),
        (
            f"{report_name}.docx",
            REPORT_CONTENT_TYPES["docx"],
            build_docx(
                title="年度财务报表审计报告（草稿，未经签发）",
                report_text=artifact_report_text,
            ),
            "annual_report_docx",
            report_version,
        ),
        (
            f"{report_name}.xlsx",
            REPORT_CONTENT_TYPES["xlsx"],
            build_report_xlsx(
                report_text=artifact_report_text,
                snapshot=snapshot,
                report_version=report_version,
            ),
            "annual_report_xlsx",
            report_version,
        ),
    ]
    for workpaper in workpapers:
        code = str(workpaper.get("code") or "workpaper")
        version = int(workpaper.get("version") or 1)
        facts = (snapshot.get("workpaper_facts") or {}).get(code) or {}
        payloads.append(
            (
                f"workpaper-{code}-{engagement_id}-v{version}.xlsx",
                REPORT_CONTENT_TYPES["xlsx"],
                build_workpaper_xlsx(
                    code=code,
                    name=str(workpaper.get("name") or code),
                    facts=facts,
                    version=version,
                ),
                f"workpaper_{code}",
                version,
            )
        )

    published: list[ArtifactRef] = []
    errors: list[str] = []
    for file_name, content_type, data, artifact_type, version in payloads:
        try:
            uploaded = service.upload_artifact(
                project_id=engagement_id,
                file_name=file_name,
                content_type=content_type,
                file_bytes=data,
            )
            published.append(
                ArtifactRef(
                    artifact_type=artifact_type,
                    template_version=str(snapshot.get("report_template_version") or ""),
                    version=version,
                    storage_ref=uploaded.storage_ref,
                    content_type=content_type,
                    file_name=file_name,
                    status="draft",
                )
            )
        except Exception as exc:
            errors.append(f"{file_name}: {str(exc)[:240]}")

    return {
        "status": "draft_saved" if not errors else "partial",
        "artifacts": [item.to_dict() for item in published],
        "errors": errors,
    }


__all__ = [
    "build_docx",
    "build_report_xlsx",
    "build_workpaper_xlsx",
    "publish_annual_artifacts",
]
