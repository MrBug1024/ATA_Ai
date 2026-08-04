from pathlib import Path
try:
    import openpyxl
    print("openpyxl", openpyxl.__version__)
except Exception as e:
    print("openpyxl fail", e)
    openpyxl = None
try:
    import pandas as pd
    print("pandas", pd.__version__)
except Exception as e:
    print("pandas fail", e)
    pd = None

xlsx = Path(r"E:\实验室\NpaLang\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】\7、案例\北京有限公司-京创会审字[2024]第3999号-标准无保留意见-4.3.V2\北京有限公司2023年年审底稿.xlsx")
xls = Path(r"E:\实验室\NpaLang\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】\7、案例\北京有限公司-京创会审字[2024]第3999号-标准无保留意见-4.3.V2\报告\2、经审计的财务报表-（会计准则）.xls")
print("xlsx", xlsx.exists(), xlsx.stat().st_size if xlsx.exists() else None)
print("xls", xls.exists(), xls.stat().st_size if xls.exists() else None)
if openpyxl and xlsx.exists():
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    print("SHEET_COUNT", len(wb.sheetnames))
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        sample = []
        for r in ws.iter_rows(max_row=5, values_only=True):
            vals = [str(c)[:40] if c is not None else "" for c in (list(r)[:8] if r else [])]
            if any(str(v).strip() for v in vals):
                sample.append(" | ".join(vals))
            if len(sample) >= 2:
                break
        print(f"{i+1:03d}\t{name}\t{sample}")
    wb.close()
if pd is not None and xls.exists():
    try:
        xl = pd.ExcelFile(xls)
        print("XLS_SHEETS", xl.sheet_names)
        for name in xl.sheet_names:
            df = pd.read_excel(xls, sheet_name=name, header=None)
            print(" ", name, df.shape)
    except Exception as e:
        print("xls err", type(e).__name__, e)
