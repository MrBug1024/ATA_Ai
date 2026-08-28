import base64
import json
from contextlib import contextmanager
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from ai_hunter.annual_audit import import_service
from ai_hunter.annual_audit.import_service import (
    TabularSheet,
    bind_structured_source_refs,
    detect_sheet_schema,
    is_audit_workpaper_workbook,
    normalize_sheet_rows,
)
from ai_hunter.annual_audit.workpaper_case import summarize_workpaper_sheets
from ai_hunter.app.graph.nodes.load_chunks import _build_chunk_batch_pages
from ai_hunter.app.subgraphs.ingest_graph import (
    _build_file_cache_key,
    _extract_spreadsheet_with_layout,
    _spreadsheet_layout_counts,
)


def _normalize(rows, *, name="Sheet1", datemode=None, period_end=date(2025, 12, 31)):
    return normalize_sheet_rows(
        TabularSheet(name=name, rows=rows, excel_datemode=datemode),
        source_ref="minio://raw/case-1/source.xlsx",
        file_name="source.xlsx",
        default_period_end=period_end,
    )


def test_receivable_sheet_is_detected_and_keeps_row_locator():
    rows = [
        ["应收账款明细"],
        ["客户名称", "单据号", "发生日期", "到期日", "期末余额", "是否关联方"],
        ["甲客户", "INV-1", "2025-01-02", "2025-03-01", "1,234.50", "是"],
    ]

    detected = detect_sheet_schema(rows, source_hint="应收明细")
    assert detected is not None
    assert detected[0] == "receivable_item"

    dataset, normalized = _normalize(rows)
    assert dataset == "receivable_item"
    assert normalized[0]["customer_name"] == "甲客户"
    assert normalized[0]["balance"] == Decimal("1234.50")
    assert normalized[0]["is_related_party"] is True
    assert normalized[0]["source_locator_json"]["sheet_name"] == "Sheet1"
    assert normalized[0]["source_locator_json"]["row_number"] == 3


def test_bank_sheet_uses_separate_inflow_and_outflow_columns():
    rows = [
        ["交易日期", "银行账号", "收入金额", "支出金额", "对方户名", "流水号"],
        ["2025/12/30", "6222", 800000, None, "甲公司", "T-1"],
        ["2025/12/31", "6222", None, 250000, "乙公司", "T-2"],
    ]

    dataset, normalized = _normalize(rows, name="银行流水")
    assert dataset == "bank_transaction"
    assert [(item["amount"], item["direction"]) for item in normalized] == [
        (Decimal("800000"), "in"),
        (Decimal("250000"), "out"),
    ]


def test_journal_sheet_normalizes_voucher_lines():
    rows = [
        ["凭证日期", "凭证号", "分录号", "科目编码", "科目名称", "借方金额", "贷方金额", "摘要"],
        ["2025-06-30", "记-100", 1, "1122", "应收账款", 1000, 0, "确认收入"],
        ["2025-06-30", "记-100", 2, "6001", "主营业务收入", 0, 1000, "确认收入"],
    ]

    dataset, normalized = _normalize(rows, name="序时账")
    assert dataset == "journal_entry"
    assert normalized[1]["voucher_no"] == "记-100"
    assert normalized[1]["line_no"] == 2
    assert normalized[1]["credit_amount"] == Decimal("1000")


def test_split_payroll_sheet_accepts_voucher_number_count_header():
    rows = [
        ["日期", "凭证号数", "科目编码", "科目名称", "摘要", "借方金额", "贷方金额", "对方科目"],
        [44946, "记-0360", "550103", "工资", "付员工工资", 1000, 0, "1002银行存款"],
    ]

    dataset, normalized = _normalize(rows, name="工资", datemode=0)

    assert dataset == "journal_entry"
    assert normalized[0]["voucher_no"] == "记-0360"
    assert normalized[0]["account_name"] == "工资"
    assert normalized[0]["debit_amount"] == Decimal("1000")


def test_split_receivable_workpaper_realigns_merged_customer_header():
    rows = [
        ["序", "往来单位", None, "业务内容", "期初余额", None, "本期借方", "本期贷方", "期末未审数", None, "调整数", None, "审定数"],
        ["号", "简称", None, "摘要", "原币", "记账本位币", "发生额", "发生额", "原币", "记账本位币"],
        [1, None, "甲客户", None, None, 3000, None, 3000, None, 0, None, None, 0],
        [2, None, "乙客户", None, None, 0, 24000, 0, None, 24000, None, None, 24000],
    ]

    detected = detect_sheet_schema(rows, source_hint="C5-2应收帐款审定表")
    assert detected is not None
    assert detected[0] == "receivable_item"
    assert detected[2]["customer_name"] == 2

    dataset, normalized = _normalize(rows, name="C5-2应收帐款审定表")
    assert dataset == "receivable_item"
    assert [item["customer_name"] for item in normalized] == ["甲客户", "乙客户"]
    assert normalized[1]["balance"] == Decimal("24000")


def test_account_balance_defaults_period_end_from_engagement():
    rows = [
        ["科目编码", "科目名称", "本期借方发生额", "本期贷方发生额", "期末借方余额", "期末贷方余额"],
        ["1002", "银行存款", 5000, 2000, 3000, 0],
    ]

    dataset, normalized = _normalize(rows, name="科目余额表")
    assert dataset == "account_balance"
    assert normalized[0]["period_end"] == date(2025, 12, 31)
    assert normalized[0]["closing_debit"] == Decimal("3000")


def test_xls_serial_date_uses_workbook_datemode():
    rows = [
        ["客户名称", "到期日", "期末余额"],
        ["甲客户", 45658, 100],
    ]

    dataset, normalized = _normalize(rows, name="应收账款", datemode=0)
    assert dataset == "receivable_item"
    assert normalized[0]["due_date"] == date(2025, 1, 1)


def test_audit_workpaper_pack_is_not_treated_as_source_ledger():
    sheets = [
        TabularSheet(name="底稿目录", rows=[]),
        TabularSheet(name="审计程序", rows=[]),
        TabularSheet(name="审定表", rows=[]),
        TabularSheet(
            name="截止测试",
            rows=[
                ["交易日期", "金额", "对方单位"],
                ["2025-12-31", 1000000, "甲公司"],
            ],
        ),
    ]

    assert is_audit_workpaper_workbook(sheets) is True


def test_plain_accounting_export_is_not_mistaken_for_workpaper():
    sheets = [
        TabularSheet(
            name="银行流水",
            rows=[
                ["交易日期", "金额", "对方单位"],
                ["2025-12-31", 1000000, "甲公司"],
            ],
        )
    ]

    assert is_audit_workpaper_workbook(sheets) is False


def test_complete_case_workpaper_pack_is_indexed_without_becoming_raw_ledger():
    sheets = [
        TabularSheet(name="封皮", rows=[["京创会审字[2024]第3999号"]]),
        TabularSheet(name="C1-2货币资金审定", rows=[["审计结论："], ["√", 2521324.82]]),
        TabularSheet(name="F1-2主营业务收入审定表", rows=[["审计结论："], ["合计", 48036043.11]]),
        TabularSheet(name="三级复核", rows=[["审计工作底稿三级复核记录"]]),
    ] + [TabularSheet(name=f"C9-{index}", rows=[["项目", index]]) for index in range(1, 98)]

    summary = summarize_workpaper_sheets(sheets, file_name="北京有限公司2023年年审底稿.xlsx")

    assert summary["is_complete_case"] is True
    assert summary["case_pack_type"] == "full_annual_audit_case"
    assert summary["sheet_count"] == 101
    assert "financial_statements" not in summary["covered_categories"]
    assert summary["program_evidence"]["C1"]["has_evidence"] is True
    assert summary["program_evidence"]["F1"]["has_conclusion"] is True


def test_spreadsheet_evidence_extraction_reads_every_sheet_without_remote_ocr():
    workbook = Workbook()
    first = workbook.active
    first.title = "封皮"
    first.append(["项目", "北京有限公司"])
    second = workbook.create_sheet("科目余额表")
    second.append(["科目编码", "科目名称", "期末余额"])
    second.append(["1002", "银行存款", 3000])
    third = workbook.create_sheet("空白底稿")
    buffer = BytesIO()
    workbook.save(buffer)

    result = _extract_spreadsheet_with_layout(
        {
            "name": "三工作表底稿.xlsx",
            "extension": ".xlsx",
            "content": base64.b64encode(buffer.getvalue()).decode("ascii"),
        }
    )

    assert result["message"] == "spreadsheet-local-read"
    assert result["raw_response"]["sheet_count"] == 3
    assert [page["sheet_name"] for page in result["pages"]] == ["封皮", "科目余额表", "空白底稿"]
    assert {block["page_idx"] for block in result["blocks"]} == {0, 1, 2}
    assert "## 工作表：科目余额表" in result["text"]
    assert "[第2行] 1002 | 银行存款 | 3000" in result["text"]
    assert "## 工作表：空白底稿\n（空工作表）" in result["text"]
    assert _spreadsheet_layout_counts(
        {"ocr_layout_results": [{"layout_result": result}]}
    ) == (1, 3, 3)


def test_storage_backed_file_cache_key_matches_persisted_sha256():
    assert _build_file_cache_key(
        {
            "name": "底稿.xlsx",
            "storage_ref": "minio://annual-raw/case-1/workbook.xlsx",
            "file_hash": "abc123",
        }
    ) == "abc123"


class _BindingMysqlCursor:
    def __init__(self, *, rows_by_table):
        self.rows_by_table = rows_by_table
        self.calls = []
        self._rows = []

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, query, params=None):
        self.calls.append((query, params))
        if "SELECT id, source_locator_json" not in query:
            return
        self._rows = []
        for table_name, rows in self.rows_by_table.items():
            if table_name in query:
                self._rows = list(rows)
                return

    def fetchall(self):
        return list(self._rows)


class _BindingPostgresCursor:
    def __init__(self):
        self.calls = []

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, query, params=None):
        self.calls.append((query, params))

    def fetchone(self):
        return {"page_image_ref": "", "page_width": 0, "page_height": 0}


class _BindingConnection:
    def __init__(self, cursor):
        self.cursor_value = cursor
        self.committed = False

    def cursor(self):
        return self.cursor_value

    def commit(self):
        self.committed = True


def _patch_binding_connections(monkeypatch, *, mysql_cursor, postgres_cursor):
    mysql_connection = _BindingConnection(mysql_cursor)
    postgres_connection = _BindingConnection(postgres_cursor)

    @contextmanager
    def fake_mysql_connection(_settings):
        yield mysql_connection

    @contextmanager
    def fake_postgres_connection(_settings):
        yield postgres_connection

    monkeypatch.setattr(import_service, "mysql_connection", fake_mysql_connection)
    monkeypatch.setattr(import_service, "postgres_connection", fake_postgres_connection)
    return mysql_connection, postgres_connection


def _binding_batch(*, chunks):
    return {
        "files": [
            {
                "id": 51,
                "file_name": "银行流水.xlsx",
                "file_sha256": "bank-file-sha",
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "storage_ref": "minio://annual/raw/bank.xlsx",
            }
        ],
        "pages": [
            {
                "id": 601,
                "file_id": 51,
                "page_no": 1,
                "sheet_name": "封皮",
                "page_text": "## 工作表：封皮\n[第1行] 年审资料",
            },
            {
                "id": 602,
                "file_id": 51,
                "page_no": 2,
                "sheet_name": "银行流水",
                "page_text": "## 工作表：银行流水\n[第1行] 交易日期 | 金额 | 对方户名",
            },
        ],
        "chunks": chunks,
    }


def _bank_locator():
    return {
        "source_sha256": "bank-file-sha",
        "file_name": "银行流水.xlsx",
        "sheet_name": "银行流水",
        "row_number": 2,
        "quote_text": "2025-12-31 | 100 | 甲公司",
    }


def test_chunk_batch_pages_keep_persisted_page_ids_and_worksheet_metadata():
    pages = _build_chunk_batch_pages(
        inserted_pages=[
            {"id": 601, "file_id": 51, "page_no": 1},
            {"id": 602, "file_id": 51, "page_no": 2},
        ],
        source_pages=[
            {
                "file_id": 51,
                "page_no": 1,
                "page_text": "## 工作表：封皮",
                "ocr_blocks": [{"sheet_name": "封皮"}],
            },
            {
                "file_id": 51,
                "page_no": 2,
                "page_text": "## 工作表：银行流水\n[第2行] 2025-12-31 | 100 | 甲公司",
                "ocr_blocks": [{"sheet_name": "银行流水"}],
            },
        ],
    )

    assert pages[0]["id"] == 601
    assert pages[1]["id"] == 602
    assert pages[1]["sheet_name"] == "银行流水"
    assert "[第2行]" in pages[1]["page_text"]


def test_binding_uses_matching_worksheet_and_row_chunk_not_first_page(monkeypatch):
    mysql_cursor = _BindingMysqlCursor(
        rows_by_table={
            "annual_bank_transaction": [
                {"id": 701, "source_locator_json": _bank_locator()},
            ]
        }
    )
    postgres_cursor = _BindingPostgresCursor()
    mysql_connection, postgres_connection = _patch_binding_connections(
        monkeypatch,
        mysql_cursor=mysql_cursor,
        postgres_cursor=postgres_cursor,
    )

    result = bind_structured_source_refs(
        engagement_id=10,
        settings=object(),
        chunk_batch=_binding_batch(
            chunks=[
                {
                    "chunk_id": "cover-chunk",
                    "page_id": 601,
                    "chunk_text": "## 工作表：封皮\n[第2行] 2025-12-31 | 100 | 甲公司",
                },
                {
                    "chunk_id": "bank-row-2",
                    "page_id": 602,
                    "chunk_text": (
                        "## 工作表：银行流水\n"
                        "[第1行] 交易日期 | 金额 | 对方户名\n"
                        "[第2行] 2025-12-31 | 100 | 甲公司"
                    ),
                },
            ]
        ),
    )

    updates = [(query, params) for query, params in mysql_cursor.calls if "UPDATE annual_" in query]
    assert result["status"] == "completed"
    assert result["bound_count"] == 1
    assert result["unbound_count"] == 0
    assert len(updates) == 1
    assert updates[0][1][0:3] == (51, 602, "bank-row-2")
    locator = json.loads(updates[0][1][4])
    assert locator["row_start"] == 2
    assert locator["row_end"] == 2
    assert locator["cell_range"] == "A2:XFD2"
    assert locator["source_chunk_row_start"] == 1
    assert locator["source_chunk_row_end"] == 2
    assert mysql_connection.committed is True
    assert postgres_connection.committed is True


def test_binding_never_marks_a_row_bound_without_a_canonical_chunk(monkeypatch):
    mysql_cursor = _BindingMysqlCursor(
        rows_by_table={
            "annual_bank_transaction": [
                {"id": 702, "source_locator_json": _bank_locator()},
            ]
        }
    )
    postgres_cursor = _BindingPostgresCursor()
    _patch_binding_connections(
        monkeypatch,
        mysql_cursor=mysql_cursor,
        postgres_cursor=postgres_cursor,
    )

    result = bind_structured_source_refs(
        engagement_id=10,
        settings=object(),
        chunk_batch=_binding_batch(chunks=[]),
    )

    updates = [query for query, _params in mysql_cursor.calls if "UPDATE annual_" in query]
    assert result["status"] == "partial"
    assert result["bound_count"] == 0
    assert result["unbound_count"] == 1
    assert result["unbound_reasons"] == {"chunk_not_found": 1}
    assert result["unbound_rows"] == [
        {
            "dataset": "bank_transaction",
            "record_id": 702,
            "reason": "chunk_not_found",
            "file_name": "银行流水.xlsx",
            "sheet_name": "银行流水",
            "row_number": 2,
        }
    ]
    assert updates == []


def test_binding_never_falls_back_to_the_first_page_when_sheet_metadata_is_missing(monkeypatch):
    mysql_cursor = _BindingMysqlCursor(
        rows_by_table={
            "annual_bank_transaction": [
                {"id": 703, "source_locator_json": _bank_locator()},
            ]
        }
    )
    postgres_cursor = _BindingPostgresCursor()
    _patch_binding_connections(
        monkeypatch,
        mysql_cursor=mysql_cursor,
        postgres_cursor=postgres_cursor,
    )
    batch = _binding_batch(
        chunks=[
            {
                "chunk_id": "first-page-chunk",
                "page_id": 601,
                "chunk_text": "[第2行] 2025-12-31 | 100 | 甲公司",
            }
        ]
    )
    # This represents an old/incomplete heavy batch.  It has database page IDs
    # but no worksheet metadata, so binding it to page 1 would be a fabrication.
    batch["pages"] = [
        {"id": 601, "file_id": 51, "page_no": 1},
        {"id": 602, "file_id": 51, "page_no": 2},
    ]

    result = bind_structured_source_refs(
        engagement_id=10,
        settings=object(),
        chunk_batch=batch,
    )

    updates = [query for query, _params in mysql_cursor.calls if "UPDATE annual_" in query]
    assert result["bound_count"] == 0
    assert result["unbound_reasons"] == {"sheet_not_found": 1}
    assert updates == []


def test_binding_refuses_ambiguous_chunks_without_writing_a_partial_anchor(monkeypatch):
    mysql_cursor = _BindingMysqlCursor(
        rows_by_table={
            "annual_bank_transaction": [
                {"id": 704, "source_locator_json": _bank_locator()},
            ]
        }
    )
    postgres_cursor = _BindingPostgresCursor()
    _patch_binding_connections(
        monkeypatch,
        mysql_cursor=mysql_cursor,
        postgres_cursor=postgres_cursor,
    )

    result = bind_structured_source_refs(
        engagement_id=10,
        settings=object(),
        chunk_batch=_binding_batch(
            chunks=[
                {
                    "chunk_id": "bank-row-2-a",
                    "page_id": 602,
                    "chunk_text": "[第2行] 2025-12-31 | 100 | 甲公司",
                },
                {
                    "chunk_id": "bank-row-2-b",
                    "page_id": 602,
                    "chunk_text": "[第2行] 2025-12-31 | 100 | 甲公司",
                },
            ]
        ),
    )

    updates = [query for query, _params in mysql_cursor.calls if "UPDATE annual_" in query]
    assert result["status"] == "partial"
    assert result["bound_count"] == 0
    assert result["unbound_reasons"] == {"row_range_ambiguous": 1}
    assert updates == []


def test_binding_refuses_ambiguous_same_name_source_files(monkeypatch):
    mysql_cursor = _BindingMysqlCursor(
        rows_by_table={
            "annual_bank_transaction": [
                {
                    "id": 705,
                    "source_locator_json": {
                        **_bank_locator(),
                        "source_sha256": "",
                    },
                },
            ]
        }
    )
    postgres_cursor = _BindingPostgresCursor()
    _patch_binding_connections(
        monkeypatch,
        mysql_cursor=mysql_cursor,
        postgres_cursor=postgres_cursor,
    )
    batch = _binding_batch(chunks=[])
    batch["files"].append(
        {
            **batch["files"][0],
            "id": 52,
            "file_sha256": "another-bank-file-sha",
        }
    )

    result = bind_structured_source_refs(
        engagement_id=10,
        settings=object(),
        chunk_batch=batch,
    )

    updates = [query for query, _params in mysql_cursor.calls if "UPDATE annual_" in query]
    assert result["status"] == "partial"
    assert result["bound_count"] == 0
    assert result["unbound_reasons"] == {"source_file_ambiguous": 1}
    assert updates == []
