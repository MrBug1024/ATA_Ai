import io
import zipfile

from openpyxl import load_workbook

from ai_hunter.annual_audit.artifact_service import (
    build_docx,
    build_report_xlsx,
    build_workpaper_xlsx,
)


def test_report_and_workpaper_artifacts_are_openable():
    report = build_report_xlsx(
        report_text="# annual audit",
        snapshot={
            "report_template_version": "report-v1",
            "readiness": {"counts": {"bank_transaction_rows": 2}},
            "sales_receivables": {"findings": []},
            "cash_and_bank": {"findings": []},
        },
        report_version=1,
    )
    workpaper = build_workpaper_xlsx(
        code="C1-2",
        name="Cash and bank",
        facts={"row_count": 2},
        version=1,
    )

    assert load_workbook(io.BytesIO(report), read_only=True).sheetnames == [
        "审计报告",
        "审计发现",
        "资料完整性",
    ]
    assert load_workbook(io.BytesIO(workpaper), read_only=True).sheetnames == ["底稿"]


def test_docx_artifact_has_required_package_parts():
    document = build_docx(title="Annual audit", report_text="Draft")

    with zipfile.ZipFile(io.BytesIO(document)) as archive:
        assert "[Content_Types].xml" in archive.namelist()
        assert "_rels/.rels" in archive.namelist()
        assert "word/document.xml" in archive.namelist()
